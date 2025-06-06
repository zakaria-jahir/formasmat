import csv
from email.utils import parsedate
from django.urls import path
from django.contrib.auth import views as auth_views
from rest_framework.response import Response
from core.serializers import CompletedTrainingSerializer, SessionSerializer, TrainingWishSerializer
from . import views
from django.views.decorators.http import require_POST
from io import BytesIO
from django.db.models import Prefetch,Case, When, Value, IntegerField
from django.apps import apps
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.db.models import Q, F
from django.utils import timezone
from django.db import IntegrityError
from django.urls import reverse
from django.views.decorators.http import require_http_methods,require_GET   
from django.contrib.auth.forms import AuthenticationForm
from django.utils.timezone import now
from datetime import timedelta
from django.http import HttpResponse, HttpResponseNotAllowed, JsonResponse




from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.platypus import  Paragraph, SimpleDocTemplate, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet,ParagraphStyle
from django.db import transaction
from django.views.decorators.csrf import csrf_exempt
from django.utils.encoding import smart_str
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from core.utils import ajax_login_required, get_coordinates_from_postal_code, haversine1

from .models import (
    TrainingRoomComment, User, Formation, Trainer, TrainingRoom, TrainingWish, Session, 
    SessionDate, Participant, SessionParticipant, ParticipantComment, 
    Notification, CompletedTraining
)
from .forms import (
    UserRegistrationForm, UserProfileForm, FormationForm, 
    TrainingRoomForm, TrainerForm, TrainingWishForm, 
    SessionForm, CustomSessionForm
)

from django.http import JsonResponse
from django.contrib.auth import get_user_model
from rest_framework.decorators import api_view

import json
import random
from geopy.distance import geodesic
import logging
import pandas as pd
import os

logger = logging.getLogger(__name__)

@login_required
@api_view(['GET'])
def get_user_info(request):
    user = request.user
    sessions = Session.objects.filter(session_participants__user=user).distinct()
    wishes = TrainingWish.objects.filter(user=user).distinct()
    completed_trainings = CompletedTraining.objects.filter(user=user).distinct()

    sessions_data = SessionSerializer(sessions, many=True).data
    wishes_data = TrainingWishSerializer(wishes, many=True).data
    completed_trainings_data = CompletedTrainingSerializer(completed_trainings, many=True).data

    return Response({
        "id": user.id,
        "email": user.email,
        "first_name": user.first_name,
        "last_name": user.last_name,
        "sessions": sessions_data,
        "wishes": wishes_data,
        "completed_trainings": completed_trainings_data,
    })
@csrf_exempt
def api_login(request):
    if request.method == 'POST':
        import json
        data = json.loads(request.body)
        email = data.get('email')
        password = data.get('password')

        user = authenticate(request, username=email, password=password)
        if user is not None:
            login(request, user)
            return JsonResponse({'success': True, 'message': 'Connexion réussie', 'user_id': user.id})
        else:
            return JsonResponse({'success': False, 'message': 'Identifiants invalides'}, status=401)
    return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
@csrf_exempt
def api_register(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            username = data.get('username')
            first_name = data.get('firstName')
            last_name = data.get('lastName')
            email = data.get('email')
            password = data.get('password1')

            if User.objects.filter(username=username).exists():
                return JsonResponse({'success': False, 'message': 'Ce nom d’utilisateur existe déjà.'}, status=400)

            user = User.objects.create_user(
                username=username,
                password=password,
                email=email,
                first_name=first_name,
                last_name=last_name
            )
            return JsonResponse({'success': True, 'message': 'Utilisateur créé avec succès.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    return JsonResponse({'error': 'Méthode non autorisée'}, status=405)

def home(request):
    """Page d'accueil du site."""
    if request.user.is_authenticated:
        if request.user.is_staff:
            return redirect('core:manage_session')
        else:
            return redirect('core:formation_list')
    return render(request, 'core/home.html')

def login_view(request):
    """Vue de connexion."""
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('core:home')
    else:
        form = AuthenticationForm()
    
    # Ajout des classes Bootstrap aux champs du formulaire
    for field in form.fields.values():
        field.widget.attrs['class'] = 'form-control'
    
    return render(request, 'core/login.html', {'form': form})

def logout_view(request):
    logout(request)
    return redirect('core:home')

def register(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)

            # Géolocalisation via code postal uniquement
            code_postal = form.cleaned_data.get('code_postal')
            city = form.cleaned_data.get('city')
            print(f"Tentative de géolocalisation pour CP: {code_postal}, Ville: {city}")

            lat, lng = get_coordinates_from_postal_code(code_postal, city)
            print(f"Coordonnées obtenues : lat={lat}, lng={lng}")

            if lat is None or lng is None:
                print("⚠️ Échec de la géolocalisation, lat/lon sont None")

            user.latitude = lat
            user.longitude = lng

            try:
                user.save()
                print(f"Utilisateur {user.get_full_name()} enregistré avec succès.")
            except Exception as e:
                print(f"❌ Erreur lors de la sauvegarde de l'utilisateur : {e}")
            
            login(request, user)
            messages.success(request, 'Inscription réussie !')
            return redirect('core:home')
        else:
            print(f"❌ Formulaire invalide : {form.errors}")
    else:
        form = UserRegistrationForm()
    
    return render(request, 'core/register.html', {'form': form})


    # check email && username in database
User = get_user_model()
def check_user_existence(request):
    """Vérifie si l'email ou le nom d'utilisateur existe déjà."""
    username = request.GET.get('username', None)
    email = request.GET.get('email', None)

    response = {
        'username_exists': False,
        'email_exists': False,
    }

    if username and User.objects.filter(username=username).exists():
        response['username_exists'] = True

    if email and User.objects.filter(email=email).exists():
        response['email_exists'] = True

    return JsonResponse(response)

@login_required

def profile(request):
    if request.method == 'POST':
        form = UserProfileForm(request.POST, instance=request.user)
        if form.is_valid():
            user = form.save(commit=False)
            if not form.cleaned_data['rpe'] and form.cleaned_data['other_rpe']:
                user.other_rpe = form.cleaned_data['other_rpe']
            else:
                user.other_rpe = ''
            user.save()
            messages.success(request, 'Profil mis à jour avec succès.')
            return redirect('core:profile')
    else:
        form = UserProfileForm(instance=request.user)

    # Sessions à venir
    upcoming_sessions = Session.objects.filter(
        session_participants__user=request.user,
        session_participants__status__in=['CONTACTED', 'FILE_EMAIL', 'FILE_PAPER']
    ).select_related('formation').prefetch_related('dates', 'dates__location', 'session_participants').distinct()

    # Souhaits de formation
    training_wishes_with_session = TrainingWish.objects.filter(
        user=request.user,
        session__isnull=False
    ).select_related('formation', 'session').order_by('-created_at')

    training_wishes_without_session = TrainingWish.objects.filter(
        user=request.user,
        session__isnull=True
    ).select_related('formation').order_by('-created_at')

    # Formations complétées via CompletedTraining
    completed_trainings = list(CompletedTraining.objects.filter(
        user=request.user
    ).select_related('formation').order_by('-completion_date'))

    # ➕ Formations complétées via sessions marquées "TERMINEE"
    completed_session_participations = SessionParticipant.objects.filter(
        user=request.user,
        session__status='TERMINEE'
    ).select_related('session__formation')

    for sp in completed_session_participations:
        completed_trainings.append({
            'formation': sp.session.formation,
            'completion_date': sp.session.end_date,
            'certificate_number': getattr(sp, 'certificate_number', None)
        })

    context = {
        'form': form,
        'upcoming_sessions': upcoming_sessions,
        'training_wishes_with_session': training_wishes_with_session,
        'training_wishes_without_session': training_wishes_without_session,
        'completed_trainings': completed_trainings,
    }
    return render(request, 'core/profile.html', context)
@login_required
@staff_member_required
@csrf_exempt
def update_session(request, session_id):
    """Handles session editing via AJAX."""
    if request.method == 'POST':
        try:
            session = get_object_or_404(Session, id=session_id)
            formation_id = request.POST.get('formation')
            trainers = request.POST.getlist('trainers[]')
            status = request.POST.get('status')
            session_dates = request.POST.getlist('session_dates[]')
            session_rooms = request.POST.getlist('session_rooms[]')
            iperia_opening = request.POST.get('iperia_opening')
            iperia_deadline = request.POST.get('iperia_deadline')

            # Validate formation
            formation = get_object_or_404(Formation, id=formation_id)

            # Update session
            with transaction.atomic():
                session.formation = formation
                session.status = status
                session.iperia_opening_date = iperia_opening
                session.iperia_deadline = iperia_deadline
                session.save()

                # Update trainers
                session.trainers.clear()
                for trainer_id in trainers:
                    trainer = get_object_or_404(Trainer, id=trainer_id)
                    session.trainers.add(trainer)

                # Update session dates and rooms
                session.dates.all().delete()  # Clear all existing dates
                for date, room_id in zip(session_dates, session_rooms):
                    if date:  # Ensure the date is not empty
                        room = get_object_or_404(TrainingRoom, id=room_id) if room_id else None
                        SessionDate.objects.create(session=session, date=date, location=room)

            return JsonResponse({'success': True, 'message': 'Session updated successfully.'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request method.'})

@login_required
def user_wishes(request):
    """Liste des souhaits de formation de l'utilisateur."""
    # Récupérer tous les souhaits de l'utilisateur
    wishes = TrainingWish.objects.filter(user=request.user).order_by('-created_at')
    
    # Récupérer toutes les formations actives
    formations = Formation.objects.filter(is_active=True)
    
    # Ajouter un formulaire pour créer un nouveau souhait
    if request.method == 'POST':
        form = TrainingWishForm(request.POST)
        if form.is_valid():
            # Vérifier si un souhait similaire existe déjà
            existing_wish = TrainingWish.objects.filter(
                user=request.user, 
                formation=form.cleaned_data['formation']
            ).exists()
            
            if existing_wish:
                messages.warning(request, 'Vous avez déjà un souhait pour cette formation.')
            else:
                wish = form.save(commit=False)
                wish.user = request.user
                try:
                    wish.save()
                    messages.success(request, 'Souhait de formation ajouté avec succès.')
                    return redirect('core:user_wishes')
                except IntegrityError:
                    messages.error(request, 'Vous avez déjà exprimé un souhait pour cette formation.')
    else:
        form = TrainingWishForm()
    
    context = {
        'wishes': wishes,
        'form': form,
        'formations': formations
    }
    
    return render(request, 'core/user_wishes.html', context)

@login_required
@staff_member_required
def manage_training_wishes(request):
    """Gestion des souhaits de formation."""
    wishes = TrainingWish.objects.all().order_by('-created_at')
    return render(request, 'core/manage_training_wishes.html', {
        'wishes': wishes
    })

@login_required
def delete_wish(request, pk):
    """Suppression d'un souhait de formation."""
    wish = get_object_or_404(TrainingWish, pk=pk, user=request.user)
    wish.delete()
    messages.success(request, 'Souhait supprimé avec succès.')
    return redirect('core:user_wishes')

@login_required
@staff_member_required
def get_training_wishes(request):
    """API: Liste des souhaits de formation."""
    wishes = TrainingWish.objects.all().order_by('-created_at')
    data = [{
        'id': wish.id,
        'user': wish.user.get_full_name() or wish.user.username,
        'formation': wish.formation.name,
        'status': wish.status,
        'created_at': wish.created_at.strftime('%d/%m/%Y')
    } for wish in wishes]
    return JsonResponse(data, safe=False)

from django.views.decorators.http import require_POST

@require_POST
@login_required
@staff_member_required
def assign_to_session(request):
    """API: Assigner un souhait à une session."""
    if request.method == 'POST':
        wish_id = request.POST.get('wish_id')
        session_id = request.POST.get('session_id')
        
        wish = get_object_or_404(TrainingWish, pk=wish_id)
        session = get_object_or_404(Session, pk=session_id)
        
        # Créer un participant
        SessionParticipant.objects.create(
            session=session,
            user=wish.user,
            status='pending'
        )
        
        # Mettre à jour le statut du souhait
        wish.status = 'assigned'
        wish.save()
        
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'}, status=400)


# Formations
def formation_list_api(request):
    formations = Formation.objects.filter(is_active=True)  # facultatif : filtre les formations actives
    data = []
    for formation in formations:
        data.append({
            "id": formation.id,
            "name": formation.name,
            "code_iperia": formation.code_iperia,
            "description": formation.description,
            "duration": formation.duration,
            "type": formation.type,
            "city": formation.city,
            "code_postal": formation.code_postal,
        })
    return JsonResponse(data, safe=False)
@login_required
def formation_list(request):
    """Liste des formations disponibles, triées par proximité (puis par défaut)."""
    user = request.user

    # Base queryset
    formations_qs = Formation.objects.filter(is_active=True)

    # Filtres GET
    search_query = request.GET.get('search', '')
    type_filter = request.GET.get('type', '')
    is_presentiel = request.GET.get('is_presentiel', '') == 'true'
    is_distanciel = request.GET.get('is_distanciel', '') == 'true'
    is_asynchrone = request.GET.get('is_asynchrone', '') == 'true'
    min_duration = request.GET.get('min_duration', '')
    max_duration = request.GET.get('max_duration', '')

    # Filtres ORM
    if search_query:
        formations_qs = formations_qs.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(code_iperia__icontains=search_query)
        )
    if type_filter:
        formations_qs = formations_qs.filter(type=type_filter)
    if is_presentiel:
        formations_qs = formations_qs.filter(is_presentiel=True)
    if is_distanciel:
        formations_qs = formations_qs.filter(is_distanciel=True)
    if is_asynchrone:
        formations_qs = formations_qs.filter(is_asynchrone=True)
    if min_duration:
        formations_qs = formations_qs.filter(duration__gte=int(min_duration))
    if max_duration:
        formations_qs = formations_qs.filter(duration__lte=int(max_duration))

    # Séparation des formations
    formations_with_coords = []
    formations_without_coords = []

    for f in formations_qs:
        if f.latitude is not None and f.longitude is not None and user.latitude and user.longitude:
            distance = haversine1(user.latitude, user.longitude, f.latitude, f.longitude)
            f.distance = round(distance, 2)  # Pour affichage si besoin
            formations_with_coords.append(f)
        else:
            formations_without_coords.append(f)

    # Tri des formations avec coordonnées
    formations_with_coords.sort(key=lambda f: f.distance)

    # Fusion avec les formations sans coordonnées
    formations = formations_with_coords + formations_without_coords

    context = {
        'formations': formations,
        'search_query': search_query,
        'type_filter': type_filter,
        'min_duration': min_duration,
        'max_duration': max_duration,
        'type_choices': Formation.FORMATION_TYPES,
    }

    return render(request, 'core/formation_list.html', context)
@login_required
def formation_detail(request, pk):
    """Détails d'une formation."""
    try:
        formation = Formation.objects.get(pk=pk, is_active=True)
    except Formation.DoesNotExist:
        messages.error(request, 'Formation introuvable.')
        return redirect('core:formation_list')
    
    sessions = Session.objects.filter(formation=formation).order_by('start_date')
    return render(request, 'core/formation_detail.html', {
        'formation': formation,
        'sessions': sessions
    })

@login_required
@staff_member_required
def formation_create(request):
    """Création d'une formation avec géolocalisation."""
    if request.method == 'POST':
        form = FormationForm(request.POST, request.FILES)
        if form.is_valid():
            formation = form.save(commit=False)

            code_postal = form.cleaned_data.get('code_postal')
            city = form.cleaned_data.get('city')
            latitude, longitude = get_coordinates_from_postal_code(code_postal,city)

            formation.latitude = latitude
            formation.longitude = longitude

            try:
                formation.save()
                # Notifications (optionnel)
                messages.success(request, 'Formation créée avec succès.')
                return redirect('core:formation_detail', pk=formation.pk)
            except IntegrityError:
                messages.error(request, "Erreur lors de l'enregistrement de la formation.")

    else:
        form = FormationForm()

    return render(request, 'core/formation_form.html', {
        'form': form,
        'title': 'Nouvelle formation'
    })

@login_required
@staff_member_required
def formation_edit(request, pk):
    """Modification d'une formation."""
    formation = get_object_or_404(Formation, pk=pk)
    
    if request.method == 'POST':
        form = FormationForm(request.POST, request.FILES, instance=formation)
        if form.is_valid():
            form.save()
            messages.success(request, 'Formation modifiée avec succès.')
            return redirect('core:formation_detail', pk=pk)
    else:
        form = FormationForm(instance=formation)
    
    return render(request, 'core/formation_form.html', {
        'form': form,
        'formation': formation,
        'title': 'Modifier la formation'
    })

@login_required
@staff_member_required
def formation_delete(request, pk):
    """Suppression d'une formation."""
    formation = get_object_or_404(Formation, pk=pk)
    
    if request.method == 'POST':
        formation.delete()
        messages.success(request, 'Formation supprimée avec succès.')
        return redirect('core:formation_list')
    return redirect('core:formation_list')

# Salles
def api_training_room_detail(request, room_id):
    room = get_object_or_404(TrainingRoom, pk=room_id)
    data = {
        "id": room.id,
        "name": room.name,
        "address": room.address,
        "postal_code": room.postal_code,
        "city": room.city,
        "capacity": room.capacity,
        "equipment": room.equipment,
        "latitude": room.latitude,
        "longitude": room.longitude,
    }
    return JsonResponse(data)
def api_training_rooms(request):
    rooms = TrainingRoom.objects.all().values(
        'id', 'name', 'address', 'postal_code', 'city',
        'capacity', 'equipment', 'latitude', 'longitude'
    )
    return JsonResponse(list(rooms), safe=False)

@staff_member_required
def training_room_list(request):
    rooms = TrainingRoom.objects.all()
    return render(request, 'core/training_room_list.html', {'rooms': rooms})

@staff_member_required
def training_room_detail(request, pk):
    room = get_object_or_404(TrainingRoom, pk=pk)
    return render(request, 'core/training_room_detail.html', {'room': room})

@login_required
def training_room_create(request):
    if request.method == 'POST':
        form = TrainingRoomForm(request.POST)
        if form.is_valid():
            room = form.save()
            messages.success(request, 'Salle créée avec succès !')
            return redirect('core:training_room_detail', pk=room.pk)
    else:
        form = TrainingRoomForm()
    return render(request, 'core/training_room_form.html', {'form': form, 'title': 'Nouvelle salle'})

@login_required
def training_room_edit(request, pk):
    room = get_object_or_404(TrainingRoom, pk=pk)
    if request.method == 'POST':
        form = TrainingRoomForm(request.POST, instance=room)
        if form.is_valid():
            room = form.save()
            messages.success(request, 'Salle mise à jour avec succès !')
            return redirect('core:training_room_detail', pk=room.pk)
    else:
        form = TrainingRoomForm(instance=room)
    return render(request, 'core/training_room_form.html', {'form': form, 'title': 'Modifier la salle'})

@login_required
def training_room_delete(request, pk):
    room = get_object_or_404(TrainingRoom, pk=pk)
    if request.method == 'POST':
        room.delete()
        messages.success(request, 'Salle supprimée avec succès !')
        return redirect('core:training_room_list')
    return redirect('core:training_room_list')

# Formateurs
@login_required
def trainers_list(request):
    """Liste des formateurs."""
    trainers = Trainer.objects.all().order_by('first_name', 'last_name')
    return render(request, 'core/trainers_list.html', {
        'trainers': trainers
    })
def trainer_list_api(request):
    trainers = Trainer.objects.all()
    data = []

    for trainer in trainers:
        specialties = [formation.name for formation in trainer.specialties.all()]

        data.append({
            "id": trainer.id,
            "first_name": trainer.first_name,
            "last_name": trainer.last_name,
            "email": trainer.email,
            "phone": trainer.phone,
            "bio": trainer.bio,
            "photo": trainer.photo.url if trainer.photo else None,
            "specialties": specialties,
        })

    return JsonResponse(data, safe=False)
@login_required
def trainer_detail(request, pk):
    """Détails d'un formateur."""
    trainer = get_object_or_404(Trainer, pk=pk)
    sessions = Session.objects.filter(trainers=trainer).order_by('-start_date')
    return render(request, 'core/trainer_detail.html', {
        'trainer': trainer,
        'sessions': sessions
    })

@login_required
@staff_member_required
def trainer_create(request):
    """Création d'un formateur."""
    if request.method == 'POST':
        form = TrainerForm(request.POST, request.FILES)
        if form.is_valid():
            trainer = form.save()
            messages.success(request, f'Formateur {trainer.first_name} {trainer.last_name} créé avec succès.')
            return redirect('core:trainers_list')
    else:
        form = TrainerForm()

    return render(request, 'core/trainer_form.html', {
        'form': form,
        'title': 'Nouveau formateur'
    })

@login_required
@staff_member_required
def trainer_edit(request, pk):
    """Modification d'un formateur."""
    trainer = get_object_or_404(Trainer, pk=pk)

    if request.method == 'POST':
        form = TrainerForm(request.POST, request.FILES, instance=trainer)
        if form.is_valid():
            form.save()
            messages.success(request, 'Formateur modifié avec succès.')
            return redirect('core:trainers_list')
    else:
        form = TrainerForm(instance=trainer)

    return render(request, 'core/trainer_form.html', {
        'form': form,
        'trainer': trainer,
        'title': 'Modifier le formateur'
    })

@login_required
@staff_member_required
def trainer_delete(request, pk):
    """Suppression d'un formateur."""
    trainer = get_object_or_404(Trainer, pk=pk)
    
    if request.method == 'POST':
        trainer.delete()
        messages.success(request, 'Formateur supprimé avec succès.')
        return redirect('core:trainers_list')
    
    return render(request, 'core/trainer_confirm_delete.html', {
        'trainer': trainer
    })

# Cartes
def public_map(request):
    rooms = TrainingRoom.objects.all()
    return render(request, 'core/public_map.html', {'rooms': rooms})

@login_required
def admin_map(request):
    rooms = TrainingRoom.objects.all()
    return render(request, 'core/admin_map.html', {'rooms': rooms})

# Gestion des souhaits de formation
@staff_member_required
def manage_training_wishes(request):
    formations = Formation.objects.all()
    return render(request, 'core/manage_training_wishes.html', {
        'formations': formations
    })

@staff_member_required
def get_training_wishes(request):
    from datetime import datetime, timedelta
    from django.utils import timezone
    import logging
    logger = logging.getLogger(__name__)

    # Log des paramètres reçus
    logger.info(f"Paramètres reçus: {request.GET}")

    # Récupération des paramètres de filtrage
    formation_id = request.GET.get('formation_id')
    date_filter = request.GET.get('date_filter')
    status = request.GET.get('status')

    # Requête de base
    wishes = TrainingWish.objects.select_related('user', 'formation').all()
    logger.info(f"Nombre total de souhaits: {wishes.count()}")

    # Filtre par formation
    if formation_id:
        wishes = wishes.filter(formation_id=formation_id)
        logger.info(f"Après filtre formation: {wishes.count()} souhaits")

    # Filtre par date
    now = timezone.now()
    if date_filter == 'today':
        wishes = wishes.filter(created_at__date=now.date())
    elif date_filter == 'week':
        start_of_week = now - timedelta(days=now.weekday())
        wishes = wishes.filter(created_at__gte=start_of_week)
    elif date_filter == 'month':
        wishes = wishes.filter(created_at__year=now.year, created_at__month=now.month)
    elif date_filter == 'year':
        wishes = wishes.filter(created_at__year=now.year)
    
    if date_filter:
        logger.info(f"Après filtre date: {wishes.count()} souhaits")

    # Sérialisation des données
    data = []
    for wish in wishes:
        wish_data = {
            'id': wish.id,
            'user': {
                'id': wish.user.id,
                'username': wish.user.username,
                'full_name': f"{wish.user.first_name} {wish.user.last_name}".strip() or wish.user.username
            },
            'formation': {
                'id': wish.formation.id,
                'name': wish.formation.name
            },
            'created_at': wish.created_at.isoformat(),
            'notes': wish.notes or '',
            'status': 'assigned' if hasattr(wish, 'session') and wish.session else 'pending'
        }
        data.append(wish_data)
    
    logger.info(f"Données sérialisées: {data}")
    return JsonResponse(data, safe=False)

@staff_member_required
def get_training_wishes(request):
    from datetime import datetime, timedelta
    from django.utils import timezone
    import logging
    logger = logging.getLogger(__name__)

    # Log des paramètres reçus
    logger.info(f"Paramètres reçus: {request.GET}")

    # Récupération des paramètres de filtrage
    formation_id = request.GET.get('formation_id')
    date_filter = request.GET.get('date_filter')
    status = request.GET.get('status')

    # Requête de base
    wishes = TrainingWish.objects.select_related('user', 'formation').all()
    logger.info(f"Nombre total de souhaits: {wishes.count()}")

    # Filtre par formation
    if formation_id:
        wishes = wishes.filter(formation_id=formation_id)
        logger.info(f"Après filtre formation: {wishes.count()} souhaits")

    # Filtre par date
    now = timezone.now()
    if date_filter == 'today':
        wishes = wishes.filter(created_at__date=now.date())
    elif date_filter == 'week':
        start_of_week = now - timedelta(days=now.weekday())
        wishes = wishes.filter(created_at__gte=start_of_week)
    elif date_filter == 'month':
        wishes = wishes.filter(created_at__year=now.year, created_at__month=now.month)
    elif date_filter == 'year':
        wishes = wishes.filter(created_at__year=now.year)
    
    if date_filter:
        logger.info(f"Après filtre date: {wishes.count()} souhaits")

    # Sérialisation des données
    data = []
    for wish in wishes:
        wish_data = {
            'id': wish.id,
            'user': {
                'id': wish.user.id,
                'username': wish.user.username,
                'full_name': f"{wish.user.first_name} {wish.user.last_name}".strip() or wish.user.username
            },
            'formation': {
                'id': wish.formation.id,
                'name': wish.formation.name
            },
            'created_at': wish.created_at.isoformat(),
            'notes': wish.notes or '',
            'status': 'assigned' if hasattr(wish, 'session') and wish.session else 'pending'
        }
        data.append(wish_data)
    
    logger.info(f"Données sérialisées: {data}")
    return JsonResponse(data, safe=False)

@staff_member_required
def get_formation_rooms(request):
    formation_id = request.GET.get('formation_id')
    if not formation_id:
        return JsonResponse({'error': 'Formation non spécifiée'}, status=400)

    rooms = TrainingRoom.objects.all()
    rooms_data = [{
        'id': room.id,
        'name': room.name,
        'capacity': room.capacity,
        'address': room.address
    } for room in rooms]

    return JsonResponse({'rooms': rooms_data})

@staff_member_required
@require_http_methods(["POST"])
def assign_to_session(request):
    try:
        data = json.loads(request.body)
        session_id = data.get('session_id')
        wish_ids = data.get('wish_ids', [])

        if not session_id or not wish_ids:
            return JsonResponse({'error': 'Données manquantes'}, status=400)

        session = get_object_or_404(Session, pk=session_id)
        wishes = TrainingWish.objects.filter(id__in=wish_ids).select_related('user')

        available_seats = session.room.capacity - session.registered_users.count()
        if len(wishes) > available_seats:
            return JsonResponse({
                'error': f'Pas assez de places disponibles ({available_seats} places pour {len(wishes)} personnes)'
            }, status=400)

        for wish in wishes:
            session.registered_users.add(wish.user)
            wish.delete()

        return JsonResponse({
            'message': f'{len(wishes)} utilisateur(s) assigné(s) à la session avec succès'
        })

    except json.JSONDecodeError:
        return JsonResponse({'error': 'Données JSON invalides'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def user_wishes(request):
    """Liste des souhaits de formation de l'utilisateur."""
    # Récupérer tous les souhaits de l'utilisateur
    wishes = TrainingWish.objects.filter(user=request.user).order_by('-created_at')
    
    # Récupérer toutes les formations actives
    formations = Formation.objects.filter(is_active=True)
    
    # Ajouter un formulaire pour créer un nouveau souhait
    if request.method == 'POST':
        form = TrainingWishForm(request.POST)
        if form.is_valid():
            # Vérifier si un souhait similaire existe déjà
            existing_wish = TrainingWish.objects.filter(
                user=request.user, 
                formation=form.cleaned_data['formation']
            ).exists()
            
            if existing_wish:
                messages.warning(request, 'Vous avez déjà un souhait pour cette formation.')
            else:
                wish = form.save(commit=False)
                wish.user = request.user
                try:
                    wish.save()
                    messages.success(request, 'Votre souhait de formation a été enregistré.')
                    return redirect('core:user_wishes')
                except IntegrityError:
                    messages.error(request, 'Vous avez déjà exprimé un souhait pour cette formation.')
    else:
        form = TrainingWishForm()
    
    context = {
        'wishes': wishes,
        'form': form,
        'formations': formations
    }
    
    return render(request, 'core/user_wishes.html', context)

@login_required
def delete_wish(request, pk):
    wish = get_object_or_404(TrainingWish, pk=pk, user=request.user)
    if request.method == 'POST':
        wish.delete()
        messages.success(request, 'Votre souhait de formation a été supprimé.')
        return redirect('core:user_wishes')
    return render(request, 'core/wish_confirm_delete.html', {'wish': wish})

def session_list(request):
    # 🔍 Vérification automatique des sessions "TERMINEE" à archiver
    print("📦 Vérification des sessions à archiver...")

    for session in Session.objects.filter(status='TERMINEE', is_archive=False):
        if session.last_status_change:
            time_diff = now() - session.last_status_change
            print(f"🔍 Session ID {session.id} — Temps écoulé : {time_diff}")
            if time_diff > timedelta(minutes=2):  # ⚠️ tu voulais 2 minutes
                session.is_archive = True
                session.save()
                print(f"✅ Session ID {session.id} archivée.")

    sessions = Session.objects.filter(is_archive=False).select_related('formation')
    return render(request, 'core/sessions_list.html', {
        'sessions': sessions
    })

@require_GET
def get_formation_sessions(request):
    formation_id = request.GET.get('formation_id')

    if not formation_id:
        return JsonResponse([], safe=False)

    try:
        formation = Formation.objects.get(pk=formation_id)
    except Formation.DoesNotExist:
        return JsonResponse([], safe=False)

    sessions = Session.objects.filter(formation=formation).order_by('-start_date')

    data = [
        {
            "id": session.id,
            "start_date": session.start_date.isoformat() if session.start_date else None,
            "end_date": session.end_date.isoformat() if session.end_date else None,
            "room": session.room.name if session.room else None
        }
        for session in sessions
    ]

    return JsonResponse(data, safe=False)
@staff_member_required
def session_create(request):
    if request.method == 'POST':
        try:
            # Récupération des données de base
            formation_id = request.POST.get('formation')
            trainer_ids = request.POST.getlist('trainers[]')
            dates = request.POST.getlist('session_dates[]')
            rooms = request.POST.getlist('session_rooms[]')
            status = request.POST.get('status', 'NON_OUVERTE')
            iperia_opening = request.POST.get('iperia_opening_date')
            iperia_deadline = request.POST.get('iperia_deadline')
            comment = request.POST.get('comment')

            # Validation des données requises
            if not formation_id:
                messages.error(request, 'Formation est requise.')
                return redirect('core:formation_list')

            # Validation des dates
            if not dates:
                messages.error(request, 'Au moins une date est requise.')
                return redirect('core:formation_list')

            # Création de la session
            try:
                formation = Formation.objects.get(id=formation_id)
            except Formation.DoesNotExist:
                messages.error(request, 'Formation introuvable.')
                return redirect('core:formation_list')
            
            session = Session.objects.create(
                formation=formation,
                status=status,
                iperia_opening_date=iperia_opening or None,
                iperia_deadline=iperia_deadline or None,
                start_date=min(dates) if dates else None,
                end_date=max(dates) if dates else None
            )

            # Ajout des formateurs
            if trainer_ids:
                session.trainers.add(*trainer_ids)

            # Ajout des dates et salles
            session_dates = []
            for i, date in enumerate(dates):
                room_id = rooms[i] if i < len(rooms) and rooms[i] else None
                session_date = SessionDate.objects.create(
                    session=session,
                    date=date,
                    location_id=room_id
                )
                session_dates.append(session_date)

            # Mettre à jour start_date et end_date
            if session_dates:
                session.start_date = min(sd.date for sd in session_dates)
                session.end_date = max(sd.date for sd in session_dates)
                session.save()

            # Redirection vers la liste des sessions ou la page de détail de la session
            messages.success(request, f'Session pour {formation.name} créée avec succès.')
            path('manage-session/', views.manage_session, name='manage_session'),


        except Exception as e:
            # Log de l'erreur détaillée
            logger.error(f"Erreur lors de la création de la session : {str(e)}", exc_info=True)
            
            # Message d'erreur plus informatif
            messages.error(request, f'Erreur lors de la création de la session : {str(e)}')
            
            # Récupérer le contexte pour réafficher le formulaire
            try:
                formation = Formation.objects.get(id=formation_id)
                context = {
                    'formation': formation,
                    'trainers': Trainer.objects.all(),
                    'training_rooms': TrainingRoom.objects.all(),
                    'error_message': str(e)
                }
                return render(request, 'core/session_form.html', context)
            except:
                return redirect('core:formation_list')
    
    # Gérer la requête GET (afficher le formulaire)
    formation_id = request.GET.get('formation')
    try:
        formation = Formation.objects.get(id=formation_id)
        context = {
            'formation': formation,
            'trainers': Trainer.objects.all(),
            'training_rooms': TrainingRoom.objects.all(),
        }
        return render(request, 'core/session_form.html', context)
    except Formation.DoesNotExist:
        messages.error(request, 'Formation introuvable.')
        return redirect('core:formation_list')
    
@staff_member_required
def export_session_pdf(request, session_id):
    session = Session.objects.get(pk=session_id)
    participants = SessionParticipant.objects.filter(session=session).select_related('user')
    extra_data = {p.user_id: p for p in Participant.objects.filter(session=session)}

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    styles = getSampleStyleSheet()
    title_style = styles['Title']
    title_style.alignment = 1  # Centré
    field_style = ParagraphStyle(name='Field', fontSize=11, leading=14, spaceAfter=5)
    separator_style = ParagraphStyle(name='Separator', fontSize=10, textColor=colors.grey, spaceAfter=12)

    # Titre principal
    elements.append(Paragraph(f"Liste des participants – Session n°{session_id}", title_style))
    elements.append(Spacer(1, 10))

    # Formateurs une seule fois
    formateurs = ', '.join([f"{t.first_name} {t.last_name}" for t in session.trainers.all()]) or "Non renseigné"
    elements.append(Paragraph(f"<b>Formateurs :</b> {formateurs}", field_style))
    elements.append(Spacer(1, 15))

    # Couleurs des statuts
    color_map = {
        'Souhait': colors.lightblue,
        'Contacté': colors.skyblue,
        'Relancé': colors.lightyellow,
        'Dossier reçu par email': colors.lightgreen,
        'Dossier reçu papier': colors.green,
        'Erreur': colors.salmon,
    }

    for sp in participants:
        user = sp.user
        participant = extra_data.get(user.id)

        status_text = dict(SessionParticipant.STATUS_CHOICES).get(sp.status, 'Inconnu')
        file_status_text = dict(Participant.FILE_STATUS).get(participant.file_status, '') if participant else ''
        comments = participant.comments if participant else ''
        rpe = getattr(user, 'rpe_association', 'Non renseigné')

        # Ligne par ligne
        elements.append(Paragraph(f"<b>Nom :</b> {user.last_name}", field_style))
        elements.append(Paragraph(f"<b>Prénom :</b> {user.first_name}", field_style))
        elements.append(Paragraph(f"<b>Email :</b> {user.email}", field_style))
        elements.append(Paragraph(f"<b>RPE/Association :</b> {rpe}", field_style))

        # Statut coloré
        color = color_map.get(status_text)
        color_html = f"<font color='{color.hexval()}'>" if color else ""
        end_tag = "</font>" if color else ""
        elements.append(Paragraph(f"<b>Statut d'inscription :</b> {color_html}{status_text}{end_tag}", field_style))

        elements.append(Paragraph(f"<b>Statut Dossier :</b> {file_status_text}", field_style))
        elements.append(Paragraph(f"<b>Commentaires :</b> {comments}", field_style))
        elements.append(Spacer(1, 8))
        elements.append(Paragraph("─────────────────────────────────────────────", separator_style))

    doc.build(elements)
    buffer.seek(0)

    return HttpResponse(buffer, content_type='application/pdf', headers={
        'Content-Disposition': f'attachment; filename="session_{session_id}_participants.pdf"'
    })

from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import Session, SessionParticipant, Participant, ParticipantComment


@staff_member_required
def export_session_csv(request, session_id):
    session = Session.objects.get(pk=session_id)
    session_participants = SessionParticipant.objects.filter(session=session).select_related('user')
    participant_extra_data = {p.user_id: p for p in Participant.objects.filter(session=session)}

    # Charger les commentaires groupés par participant_id
    comments_map = {}
    for comment in ParticipantComment.objects.filter(participant__session=session).select_related('participant__user', 'author'):
        pid = comment.participant_id
        if pid not in comments_map:
            comments_map[pid] = []
        comments_map[pid].append(f"[{comment.created_at.strftime('%d/%m/%Y')} - {comment.author.get_full_name() if comment.author else 'Anonyme'}] {comment.content}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"

    # Styles
    header_fill = PatternFill(start_color="004080", end_color="004080", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    color_map = {
        'Souhait': 'ADD8E6',
        'Contacté': '87CEEB',
        'Relancé': 'FFFF99',
        'Dossier reçu par email': '90EE90',
        'Dossier reçu papier': '32CD32',
        'Erreur': 'FF9999',
    }

    # Ligne 0 : Nom de la formation
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    cell = ws.cell(row=1, column=1)
    cell.value = f"Formation : {session.formation.name}"
    cell.font = Font(bold=True)
    cell.alignment = center_alignment

    # Ligne 1 : Formateurs
    formateurs = ', '.join(f"{t.first_name} {t.last_name}" for t in session.trainers.all())
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    cell = ws.cell(row=2, column=1)
    cell.value = f"Formateurs : {formateurs or 'Aucun'}"
    cell.font = Font(bold=True)
    cell.alignment = center_alignment

    # Ligne 2 : Date et Lieu
    session_date = session.start_date.strftime("%d/%m/%Y") if session.start_date else "Date inconnue"
    session_location = f"{session.postal_code or 'CP inconnu'} {session.city or 'Ville inconnue'}"
    date_lieu_str = f"Date : {session_date} / Lieu : {session_location}"
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=7)
    cell = ws.cell(row=3, column=1)
    cell.value = date_lieu_str
    cell.font = Font(bold=True)
    cell.alignment = center_alignment

    # Ligne 3 : En-têtes
    headers = [
        "Nom", "Prénom", "Email", "RPE/Association",
        "Statut d'inscription", "Statut Dossier", "Commentaires"
    ]
    ws.append([])
    ws.append(headers)
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border

    # Lignes suivantes : Données des participants
    for sp in session_participants:
        user = sp.user
        participant = participant_extra_data.get(user.id)

        status_text = dict(SessionParticipant.STATUS_CHOICES).get(sp.status, 'Inconnu')
        file_status_text = dict(Participant.FILE_STATUS).get(participant.file_status, '') if participant else ''

        comments = "\n".join(comments_map.get(sp.id, [])) or ''

        row_data = [
            user.last_name,
            user.first_name,
            user.email,
            getattr(user, 'rpe_association', 'Non renseigné'),
            status_text,
            file_status_text,
            comments,
        ]

        ws.append(row_data)
        row_idx = ws.max_row
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = center_alignment
            cell.border = thin_border

        status_color = color_map.get(status_text)
        if status_color:
            status_cell = ws.cell(row=row_idx, column=5)
            status_cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")

    # Ajustement des colonnes
    for i, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = max_length + 2

    # Génération du fichier Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="session_{session_id}_participants.xlsx"'
    wb.save(response)
    return response
def session_list_api(request):
    sessions = Session.objects.all()

    data = []
    for session in sessions:
        data.append({
            "id": session.id,
            "title": f"{session.formation.name if session.formation else 'Sans formation'}",  # titre généré
            "start_date": session.start_date,
            "end_date": session.end_date,
            "status": session.status,
            "city": session.city,
            "address": session.address
        })

    return JsonResponse(data, safe=False)
@staff_member_required
def archive_session(request, session_id):
    if request.method == "POST":
        session = get_object_or_404(Session, pk=session_id)
        session.is_archive = True
        session.save()
        messages.success(request, "La session a été archivée.")
    return redirect('core:manage_session')  # ou la page vers laquelle tu veux revenir
@staff_member_required
def add_room_comment(request):
    if request.method == 'POST':
        room_id = request.POST.get('room_id')
        content = request.POST.get('content')
        room = get_object_or_404(TrainingRoom, pk=room_id)
        TrainingRoomComment.objects.create(
            room=room,
            content=content,
            author=request.user
        )
    return redirect('core:training_room_list')  # Redirige vers la liste des salles

@staff_member_required
def export_archived_sessions_xlsx(request):
    sessions = Session.objects.filter(is_archive=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sessions archivées"

    # Définir styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # En-têtes
    headers = ["Formation", "Date de début", "Date de fin", "Ville", "Code postal", "Statut"]
    ws.append(headers)

    for col_num, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = title
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border

    # Remplir les données
    for session in sessions:
        row = [
            session.formation.name,
            session.start_date.strftime('%d/%m/%Y') if session.start_date else '',
            session.end_date.strftime('%d/%m/%Y') if session.end_date else '',
            session.city or '',
            session.postal_code or '',
            session.get_status_display(),
        ]
        ws.append(row)
        row_idx = ws.max_row
        for col_idx in range(1, len(row)+1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = center_alignment
            cell.border = thin_border

    # Ajuster la largeur
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Générer le fichier
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sessions_archivees.xlsx"'
    wb.save(response)
    return response
@login_required
@staff_member_required
def manage_session(request):
    """Renders the session management page and archives old sessions."""

    # 🔍 Archivage automatique
    print("📦 [DEBUG] Début vérification des sessions à archiver...")

    for session in Session.objects.filter(status='TERMINEE', is_archive=False):
        if session.last_status_change:
            time_diff = now() - session.last_status_change
            print(f"🔍 Session ID {session.id} — Temps écoulé : {time_diff}")
            if time_diff > timedelta(days=30*19):
                session.is_archive = True
                session.save()
                print(f"✅ Session ID {session.id} archivée automatiquement.")
        else:
            print(f"⚠️ Session ID {session.id} n'a pas de date de changement de statut.")

    # Rendu de la page avec les données habituelles
    sessions = Session.objects.prefetch_related('trainers', 'dates', 'formation').all()
    archived_filter = request.GET.get('archived_filter', 'non')

    formations = Formation.objects.all()
    trainers = Trainer.objects.all()
    training_rooms = TrainingRoom.objects.all()
    status_choices = Session.STATUS_CHOICES
    if archived_filter == 'non':
        sessions = sessions.filter(is_archive=False)
    elif archived_filter == 'oui':
        sessions = sessions.filter(is_archive=True)
    

    return render(request, 'core/manage_session.html', {
        'sessions': sessions,
        'formations': formations,
        'trainers': trainers,
        'training_rooms': training_rooms,
        'status_choices': status_choices,
        'archived_filter': archived_filter, 
    })
@login_required
@staff_member_required
@csrf_exempt
def create_session(request):
    """Handles session creation via AJAX."""
    if request.method == 'POST':
        try:
            print("🔧 Début de la création de session")

            formation_id = request.POST.get('formation')
            trainers = request.POST.getlist('trainers[]')
            status = request.POST.get('status')
            session_dates = request.POST.getlist('session_dates[]')
            session_rooms = request.POST.getlist('session_rooms[]')
            iperia_opening = request.POST.get('iperia_opening')
            iperia_deadline = request.POST.get('iperia_deadline')

            address = request.POST.get('address')
            city = request.POST.get('city')
            postal_code = request.POST.get('postal_code')

            print(f"📥 Données reçues : formation_id={formation_id}, status={status}, address={address}, city={city}, postal_code={postal_code}")
            print(f"📅 Dates de session : {session_dates}")
            print(f"🏫 Salles de session : {session_rooms}")
            print(f"👥 Formateurs : {trainers}")

            # Validate formation
            formation = get_object_or_404(Formation, id=formation_id)
            print(f"✅ Formation trouvée : {formation.name}")

            # Create session
            with transaction.atomic():
                session = Session.objects.create(
                    formation=formation,
                    status=status,
                    iperia_opening_date=iperia_opening,
                    iperia_deadline=iperia_deadline,
                    address=address,
                    city=city,
                    postal_code=postal_code,
                )
                print(f"🆕 Session créée avec ID : {session.id}")

                # Géolocalisation automatique
                if postal_code:
                    lat, lon = get_coordinates_from_postal_code(postal_code, city)
                    print(f"📍 Coordonnées obtenues : lat={lat}, lon={lon}")
                    if lat is not None and lon is not None:
                        session.latitude = lat
                        session.longitude = lon
                        session.save()
                        print("✅ Coordonnées enregistrées dans la session")
                    else:
                        print("⚠️ Coordonnées non disponibles pour le code postal et la ville fournis")
                else:
                    print("⚠️ Code postal non fourni, géolocalisation ignorée")

                # Add trainers
                for trainer_id in trainers:
                    trainer = get_object_or_404(Trainer, id=trainer_id)
                    session.trainers.add(trainer)
                    print(f"👤 Formateur ajouté : {trainer}")

                # Add session dates and rooms
                for date, room_id in zip(session_dates, session_rooms):
                    room = get_object_or_404(TrainingRoom, id=room_id)
                    SessionDate.objects.create(session=session, date=date, location=room)
                    print(f"📆 Date ajoutée : {date} avec salle ID : {room_id}")

            print("✅ Création de session terminée avec succès")
            return JsonResponse({'success': True, 'message': 'Session créée avec succès.'})
        except Exception as e:
            print(f"❌ Erreur lors de la création de la session : {e}")
            return JsonResponse({'success': False, 'error': str(e)})
    print("⚠️ Méthode de requête invalide")
    return JsonResponse({'success': False, 'error': 'Méthode de requête invalide.'})
@staff_member_required
def assign_wishes_to_session(request, session_id):
    session = get_object_or_404(Session, pk=session_id)
    wishes = TrainingWish.objects.filter(
        formation=session.formation,
        session__isnull=True
    ).select_related('user')

    sort_by = request.GET.get('sort')
    wishes_with_distances = []

    # Récupérer les coordonnées de la session
    session_lat, session_lon = get_coordinates_from_postal_code(session.postal_code, session.city)

    if session_lat and session_lon:
        for wish in wishes:
            user = wish.user

            # ⚡ D'abord utiliser les coordonnées existantes si disponibles
            if user.latitude and user.longitude:
                user_lat, user_lon = user.latitude, user.longitude
            else:
                # Sinon essayer de récupérer via Nominatim
                user_lat, user_lon = get_coordinates_from_postal_code(user.code_postal, user.city)

            print(f"USER: {user.get_full_name()}, CP: {user.code_postal}, VILLE: {user.city}, coords: {user_lat}, {user_lon}")

            if user_lat and user_lon:
                distance = haversine1(session_lat, session_lon, user_lat, user_lon)
            else:
                distance = float('inf')  # Très loin si pas de coordonnées

            wishes_with_distances.append((wish, distance))

        if sort_by == 'distance':
            wishes_with_distances.sort(key=lambda x: x[1])
        elif sort_by == 'date':
            wishes_with_distances.sort(key=lambda x: x[0].created_at)
    else:
        # Si pas de coordonnées pour la session, on met infini pour tous
        wishes_with_distances = [(wish, float('inf')) for wish in wishes]

    context = {
        'session': session,
        'wishes_with_distances': wishes_with_distances,
    }
    return render(request, 'core/assign_wishes_to_session.html', context)
@login_required
@staff_member_required
def assign_single_wish_to_session(request, session_id, wish_id):
    session = get_object_or_404(Session, id=session_id)
    wish = get_object_or_404(TrainingWish, id=wish_id, session__isnull=True)

    wish.session = session
    wish.save()

    SessionParticipant.objects.create(
        session=session,
        user=wish.user,
        status=SessionParticipant.STATUS_WISH
    )

    messages.success(request, f"Le souhait de {wish.user.get_full_name()} a été affecté à la session.")
    return redirect('core:assign_wishes_to_session', session_id=session.id)
@login_required
@staff_member_required
def get_users(request):
    """Récupère la liste des utilisateurs pour l'ajout de participants."""
    users = User.objects.filter(is_active=True).order_by('last_name', 'first_name')
    data = [{
        'id': user.id,
        'username': user.username,
        'full_name': user.get_full_name() or user.username
    } for user in users]
    return JsonResponse(data, safe=False)

@login_required
@staff_member_required
def update_participant_status(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        participant_id = request.POST.get('participant_id')
        new_status = request.POST.get('status')
        comments = request.POST.get('comments', '')
        
        participant = Session.objects.get(id=participant_id)
        participant.status = new_status
        participant.comments = comments
        participant.save()
        return JsonResponse({
            'success': True,
            'status_display': participant.get_status_display()
        })
    except Session.DoesNotExist:
        return JsonResponse({'error': 'Participant non trouvé'}, status=404)

@login_required
@staff_member_required
def add_participant(request, session_id):
    """Vue pour ajouter un participant à une session."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        session = Session.objects.get(id=session_id)
        user_id = request.POST.get('user_id')
        user = User.objects.get(id=user_id)
        
        # Créer le participant
        participant = SessionParticipant.objects.create(
            session=session,
            user=user,
            status='CONTACTED'  # Statut par défaut
        )
        
        # Créer une notification pour l'utilisateur
        Notification.objects.create(
            user=user,
            title='Inscription à une formation',
            message=f'Vous avez été inscrit à la formation "{session.formation.name}".',
            notification_type='SESSION_REGISTRATION'
        )
        
        # Mettre à jour le voeu de formation correspondant s'il existe
        try:
            training_wish = TrainingWish.objects.get(
                user=user, 
                formation=session.formation
            )
            training_wish.session = session
            training_wish.save()
        except TrainingWish.DoesNotExist:
            # Pas de voeu trouvé, ce n'est pas une erreur bloquante
            pass
        
        return JsonResponse({
            'success': True,
            'participant': {
                'id': participant.id,
                'name': participant.user.get_full_name(),
                'email': participant.user.email,
                'status': participant.get_status_display(),
                'status_class': participant.get_status_badge_class(),
            }
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@staff_member_required
def remove_participant(request, participant_id):
    """Vue pour retirer un participant d'une session."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        # Récupérer le participant
        participant = SessionParticipant.objects.select_related('session', 'user', 'session__formation').get(id=participant_id)
        session = participant.session
        user = participant.user
        formation = session.formation

        # Créer une notification pour l'utilisateur
        Notification.objects.create(
            user=user,
            type='session_created',  # Utiliser un type existant
            message=f'Votre participation à la formation "{formation.name}" a été annulée par un administrateur.',
        )

        # Créer ou réactiver un voeu de formation
        training_wish, created = TrainingWish.objects.get_or_create(
            user=user, 
            formation=formation,
            defaults={
                'notes': 'Voeu recréé suite à annulation de participation',
                'created_at': timezone.now()
            }
        )
        
        # Toujours réinitialiser le statut du voeu
        training_wish.session = None
        training_wish.save()

        # Supprimer le participant de la session
        participant.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Participant retiré avec succès',
            'session_id': session.id
        })
    
    except SessionParticipant.DoesNotExist:
        return JsonResponse({'error': 'Participant non trouvé'}, status=404)
    except Exception as e:
        logger.error(f"Erreur lors de la suppression du participant: {e}")
        return JsonResponse({'error': str(e)}, status=400)
@require_POST
@login_required
@staff_member_required
def change_session_status(request, session_id):
    try:
        session = Session.objects.get(pk=session_id)
    except Session.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Session introuvable'}, status=404)

    new_status = request.POST.get('status')
    if new_status not in dict(Session.STATUS_CHOICES).keys():
        return JsonResponse({'success': False, 'error': 'Statut invalide'}, status=400)

    session.status = new_status
    session.save()

    return JsonResponse({
        'success': True,
        'status': session.get_status_display(),
        'status_class': session.get_status_class()  # si tu as une méthode qui retourne la bonne classe Bootstrap
    })
@login_required
@staff_member_required
def update_session_status(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        session_id = request.POST.get('session_id')
        new_status = request.POST.get('status')
        
        session = Session.objects.get(id=session_id)
        logger.info(f"Mise à jour du statut de la session {session_id} en {new_status}")
        
        # Vérifier que le statut est valide
        valid_statuses = dict(Session.STATUS_CHOICES)
        if new_status not in valid_statuses:
            logger.error(f"Statut invalide : {new_status}")
            return JsonResponse({'error': 'Statut invalide'}, status=400)
        
        # Mettre à jour le statut
        session.status = new_status
        session.save()
        logger.info(f"Statut de la session {session_id} mis à jour en {new_status}")
        

        # Créer une notification pour chaque participant
        participants = session.session_participants.all()
        for participant in participants:
            Notification.create_notification(
                user=participant.user,
                title=f'Mise à jour du statut de la formation "{session.formation.name}"',
                notification_type='session_status_update',
                message=f'Le statut a été mis à jour à "{valid_statuses[new_status]}".'
            )
            logger.info(f"Notification envoyée à {participant.user.username} pour la mise à jour du statut de la session {session_id}")

        return JsonResponse({
    'success': True,
    'message': 'Statut de la session mis à jour avec succès',
    'new_status': new_status,
    'new_status_display': valid_statuses[new_status],
    'redirect': f'/manage-session/detail/{session_id}/'
})
        
    
    except Exception as e:
        logger.error(f"Erreur lors de la mise à jour du statut de la session: {e}")
        return JsonResponse({'error': str(e)}, status=400)
def admin_training_sessions(request):
    user = request.user

    # Vérifie que l'utilisateur a bien latitude et longitude
    user_lat = getattr(user, 'latitude', None)
    user_lon = getattr(user, 'longitude', None)

    # 🔍 Vérification automatique des sessions "TERMINEE" à archiver
    print("📦 Vérification des sessions à archiver...")

    for session in Session.objects.filter(status='TERMINEE', is_archive=False):
        if session.last_status_change:
            time_diff = now() - session.last_status_change
            print(f"🔍 Session ID {session.id} — Temps écoulé : {time_diff}")
            if time_diff > timedelta(minutes=1):
                session.is_archive = True
                session.save()
                print(f"✅ Session ID {session.id} archivée.")

    formations = Formation.objects.all()
    selected_formation_id = request.GET.get('formation')
    city_filter = request.GET.get('rpe')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    # 🔒 On filtre pour n'afficher que les sessions non archivées
    sessions = Session.objects.filter(is_archive=False).select_related('formation')

    if selected_formation_id:
        sessions = sessions.filter(formation__id=selected_formation_id)

    if city_filter:
        sessions = sessions.filter(city__icontains=city_filter)

    if date_from:
        sessions = sessions.filter(start_date__gte=parsedate(date_from))
    if date_to:
        sessions = sessions.filter(end_date__lte=parsedate(date_to))

    # Ajoute la distance si latitude et longitude sont connues
    sessions_with_distance = []
    for session in sessions:
        if session.latitude and session.longitude and user_lat and user_lon:
            distance = haversine1(user_lat, user_lon, session.latitude, session.longitude)
        else:
            distance = float('inf')
        sessions_with_distance.append((distance, session))

    sessions_with_distance.sort(key=lambda x: x[0])
    sessions = [s for _, s in sessions_with_distance]

    context = {
        'formations': formations,
        'sessions': sessions,
        'current_formation_filter': selected_formation_id,
        'current_rpe_filter': city_filter,
        'current_date_from': date_from,
        'current_date_to': date_to,
    }
    return render(request, 'core/admin_training_sessions.html', context)
@login_required
@staff_member_required
def get_session(request, session_id):
    """Fetches session details for editing, and archive old sessions if needed."""

    print("🔍 Vérification des sessions terminées non archivées...")
    for session in Session.objects.filter(status='TERMINEE', is_archive=False):
        delta = now() - session.last_status_change
        print(f"🕓 Session ID {session.id} - Dernier changement : {session.last_status_change} (delta: {delta})")
        if session.last_status_change and delta > timedelta(minutes=1):
            session.is_archive = True
            session.save()
            print(f"✅ Session ID {session.id} archivée automatiquement.")

    session = get_object_or_404(Session, id=session_id)
    print(f"📦 Chargement de la session ID {session.id}")

    session_data = {
        'id': session.id,
        'formation_id': session.formation.id,
        'status': session.status,
        'trainers': list(session.trainers.values_list('id', flat=True)),
        'dates': [{'date': date.date, 'room_id': date.location.id if date.location else None} for date in session.dates.all()],
        'iperia_opening': session.iperia_opening_date,
        'iperia_deadline': session.iperia_deadline,
    }
    return JsonResponse(session_data)


@login_required
@staff_member_required
def delete_session(request, session_id):
    """View to delete a session."""
    if request.method == 'POST':  # Ensure the request is POST
        try:
            session = Session.objects.get(id=session_id)
            session.delete()
            return JsonResponse({'success': True, 'message': 'Session deleted successfully.'})
        except Session.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Session does not exist.'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    else:
        return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=405)

@login_required
@staff_member_required
def session_detail(request, session_id):
    """Vue détaillée d'une session."""
    logger.info(f"=== DÉBUT session_detail pour session_id={session_id} ===")
    
    try:
        # D'abord récupérer la session
        session = get_object_or_404(Session, pk=session_id)
        logger.info(f"Session trouvée : {session}")
        
        # Récupérer les participants
        participants = list(SessionParticipant.objects.filter(session=session).select_related('user'))
        logger.info(f"Participants trouvés : {len(participants)}")
        
        # Récupérer les dates
        session_dates = list(session.dates.select_related('location').all())
        logger.info(f"Dates trouvées : {len(session_dates)}")
        
        # Récupérer les utilisateurs
        participant_user_ids = [p.user.id for p in participants]
        all_users = list(User.objects.exclude(id__in=participant_user_ids))
        logger.info(f"Utilisateurs trouvés : {len(all_users)}")
        
        # Récupérer les souhaits
        participant_users = [p.user for p in participants]
        training_wishes = list(TrainingWish.objects.filter(
            formation=session.formation
        ).exclude(
            user__in=participant_users
        ).select_related('user', 'formation'))
        logger.info(f"Souhaits trouvés : {len(training_wishes)}")
        
        # Calculer les places disponibles
        total_capacity = float('inf')  # Commencer par la valeur maximale
        for date in session_dates:
            if date.location and hasattr(date.location, 'capacity') and date.location.capacity:
                total_capacity = min(total_capacity, date.location.capacity)
            else:
                total_capacity = min(total_capacity, 20)
        
        # Si aucune salle n'a été trouvée, utiliser 12 comme valeur par défaut
        if total_capacity == float('inf'):
            total_capacity = 12
        
        registered_count = len(participants)
        available_seats = max(0, total_capacity - registered_count)
        
        # Vérifier que tous les choix de statut sont disponibles
        status_choices = SessionParticipant.STATUS_CHOICES
        session_status_choices = Session.STATUS_CHOICES
        logger.info(f"Statuts participants : {status_choices}")
        logger.info(f"Statuts session : {session_status_choices}")
        
        trainers = list(session.trainers.all())
        logger.info(f"Formateurs trouvés : {len(trainers)}")
        
        # Préparer le contexte
        context = {
            'session': session,
            'participants': participants,
            'session_dates': session_dates,
            'trainers': trainers,
            'participant_status_choices': status_choices,
            'available_seats': available_seats,
            'total_capacity': total_capacity,
            'registered_count': registered_count,
            'all_users': all_users,
            'training_wishes': training_wishes,
            'session_status_choices': session_status_choices
        }
        
        logger.info("Contexte préparé avec succès")
        logger.info(f"Clés du contexte : {list(context.keys())}")
        
        try:
            # Rendre le template
            response = render(request, 'core/session_detail.html', context)
            logger.info("Template rendu avec succès")
            return response
        except Exception as template_error:
            logger.error(f"Erreur lors du rendu du template : {str(template_error)}", exc_info=True)
            messages.error(request, f"Erreur lors du rendu du template : {str(template_error)}")
            return redirect('core:manage_session')
            
    except Exception as e:
        logger.error(f"Erreur dans session_detail : {str(e)}", exc_info=True)
        messages.error(request, f"Une erreur s'est produite : {str(e)}")
        return redirect('core:manage_session')

@login_required
@staff_member_required
def add_participant_from_wish(request, session_id):
    """Ajouter un participant depuis un voeu de formation."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        session = Session.objects.get(id=session_id)
        wish_id = request.POST.get('wish_id')
        training_wish = TrainingWish.objects.get(id=wish_id)
        user = training_wish.user
        
        # Créer le participant
        participant = SessionParticipant.objects.create(
            session=session,
            user=user,
            status='CONTACTED'  # Statut par défaut
        )
        
        # Créer une notification pour l'utilisateur
        Notification.objects.create(
            user=user,
            title='Inscription à une formation',
            message=f'Vous avez été inscrit à la formation "{session.formation.name}".',
            notification_type='SESSION_REGISTRATION'
        )
        
        # Mettre à jour le voeu de formation
        training_wish.session = session
        training_wish.save()
        
        return JsonResponse({
            'success': True,
            'participant': {
                'id': participant.id,
                'name': participant.user.get_full_name(),
                'email': participant.user.email,
                'status': participant.get_status_display(),
                'status_class': participant.get_status_badge_class(),
            }
        })
    except Session.DoesNotExist:
        return JsonResponse({'error': 'Session non trouvée'}, status=404)
    except TrainingWish.DoesNotExist:
        return JsonResponse({'error': 'Voeu de formation non trouvé'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@staff_member_required
def create_and_add_participant(request, session_id):
    """Créer un nouvel utilisateur et l'ajouter à la session."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'}, status=405)
    
    try:
        session = Session.objects.get(id=session_id)
        
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        
        # Validation des champs requis
        if not all([first_name, last_name, email]):
            return JsonResponse({
                'success': False,
                'error': 'Tous les champs sont obligatoires'
            }, status=400)
        
        # Vérifier que l'email n'existe pas déjà
        if User.objects.filter(email=email).exists():
            return JsonResponse({
                'success': False,
                'error': 'Un utilisateur avec cet email existe déjà'
            }, status=400)
        
        # Créer l'utilisateur
        user = User.objects.create(
            username=email,
            first_name=first_name,
            last_name=last_name,
            email=email
        )
        user.set_unusable_password()
        user.save()
        
        # Créer le participant
        participant = SessionParticipant.objects.create(
            session=session,
            user=user,
            status='CONTACTED'  # Statut par défaut
        )
        
        return JsonResponse({
            'success': True,
            'participant': {
                'id': participant.id,
                'name': f"{participant.user.first_name} {participant.user.last_name}",
                'email': participant.user.email,
                'status': participant.get_status_display(),
                'status_class': participant.get_status_badge_class()
            }
        })
    
    except Session.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Session non trouvée'
        }, status=404)
    except Exception as e:
        logger.error(f"Erreur lors de l'ajout du participant : {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': f"Une erreur s'est produite lors de l'ajout du participant : {str(e)}"
        }, status=500)
    
    return JsonResponse({'error': 'Méthode non autorisée'}, status=405)

@csrf_exempt
@require_POST
@login_required
def api_add_training_wish(request):
    try:
        data = json.loads(request.body)  # ✅ lire le JSON envoyé
        formation_id = data.get("formation_id")
        if not formation_id:
            return JsonResponse({"error": "ID de formation manquant"}, status=400)

        formation = get_object_or_404(Formation, pk=formation_id)

        TrainingWish.objects.create(user=request.user, formation=formation, notes="")
        return JsonResponse({"success": True, "message": "Souhait ajouté avec succès !"})

    except IntegrityError:
        return JsonResponse({"success": False, "message": "Souhait déjà existant."})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)
import json
@csrf_exempt
@login_required
def api_remove_training_wish(request):
    if request.method != "POST":
        return JsonResponse({"error": "Méthode non autorisée"}, status=405)

    try:
        body = json.loads(request.body)
        formation_id = body.get("formation_id")
    except json.JSONDecodeError:
        return JsonResponse({"error": "Requête invalide"}, status=400)

    if not formation_id:
        return JsonResponse({"error": "ID de formation manquant"}, status=400)

    formation = get_object_or_404(Formation, pk=formation_id)

    try:
        wish = TrainingWish.objects.get(user=request.user, formation=formation)
        wish.delete()
        return JsonResponse({"success": True, "message": "Souhait supprimé avec succès !"})
    except TrainingWish.DoesNotExist:
        return JsonResponse({"success": False, "message": "Souhait introuvable."})
@login_required
def api_user_wishes(request):
    wishes = TrainingWish.objects.filter(user=request.user).values_list('formation_id', flat=True)
    return JsonResponse({"wished_ids": list(wishes)})
@require_POST
@login_required
def add_training_wish(request, formation_pk):
    formation = get_object_or_404(Formation, pk=formation_pk)
    try:
        TrainingWish.objects.create(user=request.user, formation=formation, notes="")
        messages.success(request, "Souhait ajouté avec succès !")
    except IntegrityError:
        messages.warning(request, "Vous avez déjà souhaité cette formation.")

    return redirect('core:formation_list')
@login_required
def notifications_list(request):
    # Ajoutez votre logique ici
    return render(request, 'core/notifications.html', {
        'notifications': [],  # Remplacez par vos notifications
    })

@staff_member_required
def admin_training_wishes(request):
    """Vue pour la gestion des souhaits de formation par les administrateurs."""
    # Récupérer tous les souhaits de formation
    wishes = TrainingWish.objects.select_related('user', 'formation').order_by('-created_at')
    
    # Filtres
    formation_filter = request.GET.get('formation', '')
    rpe_filter = request.GET.get('rpe', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Appliquer les filtres
    if formation_filter:
        wishes = wishes.filter(formation__name__icontains=formation_filter)
    
    if rpe_filter:
        wishes = wishes.filter(user__rpe_association__icontains=rpe_filter)
    
    if date_from:
        wishes = wishes.filter(created_at__date__gte=date_from)
    
    if date_to:
        wishes = wishes.filter(created_at__date__lte=date_to)
    
    # Récupérer toutes les formations pour le filtre
    all_formations = Formation.objects.all()
    
    context = {
        'wishes': wishes,
        'title': 'Gestion des souhaits de formation',
        'formations': all_formations,
        'current_formation_filter': formation_filter,
        'current_rpe_filter': rpe_filter,
        'current_date_from': date_from,
        'current_date_to': date_to,
    }
    
    return render(request, 'core/admin_training_wishes.html', context)

@staff_member_required
def assign_wish_to_session(request, wish_id):
    """Assigner un souhait de formation à une session."""
    wish = get_object_or_404(TrainingWish, id=wish_id)
    
    if request.method == 'POST':
        session_id = request.POST.get('session')
        if session_id:
            session = get_object_or_404(Session, id=session_id)
            
            # Créer un participant
            SessionParticipant.objects.create(
                session=session,
                user=wish.user,
                status='CONTACTED'  # Statut par défaut
            )
            
            # Créer une notification
            Notification.objects.create(
                user=wish.user,
                type='wish_assigned',
                message=f'Votre souhait pour la formation {wish.formation.name} a été assigné à une session',
                related_object_id=session.id,
                related_object_type='Session'
            )
            
            # Supprimer le souhait
            wish.delete()
            
            messages.success(request, 'Le souhait a été assigné à la session avec succès.')
            return redirect('core:admin_training_wishes')
    
    # Récupérer les sessions ouvertes pour cette formation
    sessions = Session.objects.filter(
        formation=wish.formation, 
        status__in=['NON_OUVERTE', 'DEMANDEE', 'OUVERTE']
    )
    
    context = {
        'wish': wish,
        'sessions': sessions,
        'title': 'Assigner un souhait à une session'
    }
    
    return render(request, 'core/assign_wish_to_session.html', context)


@login_required
@staff_member_required
def update_participant_status(request, participant_id):
    """Vue pour mettre à jour le statut d'un participant."""
    if request.method != 'POST':
        logger.error(f"Méthode non autorisée pour update_participant_status: {request.method}")
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        logger.info(f"Tentative de mise à jour du statut pour le participant {participant_id}")
        logger.info(f"Données POST reçues: {request.POST}")
        
        participant = SessionParticipant.objects.get(id=participant_id)
        new_status = request.POST.get('status')
        comment = request.POST.get('comment')
        
        logger.info(f"Nouveau statut: {new_status}")
        
        # Valider le statut
        valid_statuses = [
            SessionParticipant.STATUS_WISH,
            SessionParticipant.STATUS_CONTACTED,
            SessionParticipant.STATUS_REMINDED,
            SessionParticipant.STATUS_FILE_RECEIVED_EMAIL,
            SessionParticipant.STATUS_FILE_RECEIVED_PAPER,
            SessionParticipant.STATUS_ERROR
        ]
        
        if not new_status:
            logger.error("Aucun statut fourni")
            return JsonResponse({'error': 'Aucun statut fourni'}, status=400)
        
        if new_status not in valid_statuses:
            logger.error(f"Statut invalide: {new_status}")
            return JsonResponse({'error': 'Statut invalide'}, status=400)
        
        participant.status = new_status
        participant.save()
        
        # Ajouter un commentaire si fourni
        if comment:
            ParticipantComment.objects.create(
                participant=participant,
                author=request.user,
                content=comment
            )
        
        logger.info(f"Statut mis à jour avec succès pour le participant {participant_id}")
        
        return JsonResponse({
            'success': True,
            'status': participant.get_status_display(),
            'status_class': participant.get_status_badge_class(),
        })
    except SessionParticipant.DoesNotExist:
        logger.error(f"Participant non trouvé: {participant_id}")
        return JsonResponse({'error': 'Participant non trouvé'}, status=404)
    except Exception as e:
        logger.error(f"Erreur lors de la mise à jour du statut: {str(e)}")
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@staff_member_required
def update_participant_comments(request, participant_id):
    """Vue pour mettre à jour les commentaires d'un participant."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        participant = SessionParticipant.objects.get(id=participant_id)
        comment_content = request.POST.get('comments', '').strip()
        
        if not comment_content:
            return JsonResponse({'error': 'Le commentaire ne peut pas être vide'}, status=400)
            
        # Créer un nouveau commentaire
        ParticipantComment.objects.create(
            participant=participant,
            author=request.user,
            content=comment_content
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Commentaire ajouté avec succès'
        })
            
    except SessionParticipant.DoesNotExist:
        return JsonResponse({'error': 'Participant non trouvé'}, status=404)
    except Exception as e:
        print(f"DEBUG: Erreur lors de l'ajout du commentaire : {str(e)}")
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@staff_member_required
def get_session(request, session_id):
    """Fetches session details for editing."""
    session = get_object_or_404(Session, id=session_id)
    logger.info(f"Session data: {session}")
    session_data = {
        'id': session.id,
        'formation_id': session.formation.id,  # Ensure this is included
        'status': session.status,
        'trainers': list(session.trainers.values_list('id', flat=True)),
        'dates': [{'date': date.date, 'room_id': date.location.id if date.location else None} for date in session.dates.all()],
        'iperia_opening': session.iperia_opening_date,
        'iperia_deadline': session.iperia_deadline,
    }
    return JsonResponse(session_data)


@login_required
@staff_member_required
def delete_session(request, session_id):
    """Vue pour supprimer une session."""
    try:
        session = Session.objects.get(id=session_id)
        session.delete()
        return JsonResponse({'success': True})
    except Session.DoesNotExist:
        return JsonResponse({'error': 'Session non trouvée'}, status=404)
    except Exception as e:
        logger.error(f"Erreur lors de la suppression de la session {session_id}: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@staff_member_required
def add_participant_from_wish(request, session_id):
    """Ajouter un participant depuis un voeu de formation."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        session = Session.objects.get(id=session_id)
        wish_id = request.POST.get('wish_id')
        training_wish = TrainingWish.objects.get(id=wish_id)
        user = training_wish.user
        
        # Créer le participant
        participant = SessionParticipant.objects.create(
            session=session,
            user=user,
            status='CONTACTED'  # Statut par défaut
        )
        
        # Créer une notification pour l'utilisateur
        Notification.objects.create(
            user=user,
            title='Inscription à une formation',
            message=f'Vous avez été inscrit à la formation "{session.formation.name}".',
            notification_type='SESSION_REGISTRATION'
        )
        
        # Mettre à jour le voeu de formation
        training_wish.session = session
        training_wish.save()
        
        return JsonResponse({
            'success': True,
            'participant': {
                'id': participant.id,
                'name': participant.user.get_full_name(),
                'email': participant.user.email,
                'status': participant.get_status_display(),
                'status_class': participant.get_status_badge_class(),
            }
        })
    except Session.DoesNotExist:
        return JsonResponse({'error': 'Session non trouvée'}, status=404)
    except TrainingWish.DoesNotExist:
        return JsonResponse({'error': 'Voeu de formation non trouvé'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def add_training_wish(request, formation_pk):
    """Ajouter un souhait de formation depuis la page de détail de formation."""
    formation = get_object_or_404(Formation, pk=formation_pk)
    
    if request.method == 'POST':
        try:
            # Créer le souhait de formation
            wish = TrainingWish.objects.create(
                user=request.user,
                formation=formation,
                notes=request.POST.get('message', '')
            )
            messages.success(request, 'Votre souhait de formation a été enregistré.')
        except IntegrityError:
            messages.error(request, 'Vous avez déjà exprimé un souhait pour cette formation.')
        
        return redirect('core:formation_detail', pk=formation_pk)

@login_required
def notifications_list(request):
    """Liste des notifications de l'utilisateur."""
    # Récupérer toutes les notifications non lues
    unread_notifications = request.user.notifications.filter(is_read=False)
    
    # Récupérer les notifications lues, limitées à 20
    read_notifications = request.user.notifications.filter(is_read=True).order_by('-created_at')[:20]
    
    # Combiner et trier les notifications
    notifications = list(unread_notifications) + list(read_notifications)
    
    # Marquer les notifications comme lues
    unread_notifications.update(is_read=True)
    
    context = {
        'notifications': notifications,
        'title': 'Mes notifications'
    }
    
    return render(request, 'core/notifications_list.html', context)


@staff_member_required
def admin_training_wishes(request):
    """Vue pour la gestion des souhaits de formation par les administrateurs."""
    # Récupérer tous les souhaits de formation
    wishes = TrainingWish.objects.select_related('user', 'formation').order_by('-created_at')
    
    # Filtres
    formation_filter = request.GET.get('formation', '')
    rpe_filter = request.GET.get('rpe', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Appliquer les filtres
    if formation_filter:
        wishes = wishes.filter(formation__name__icontains=formation_filter)
    
    if rpe_filter:
        wishes = wishes.filter(user__rpe_association__icontains=rpe_filter)
    
    if date_from:
        wishes = wishes.filter(created_at__date__gte=date_from)
    
    if date_to:
        wishes = wishes.filter(created_at__date__lte=date_to)
    
    # Récupérer toutes les formations pour le filtre
    all_formations = Formation.objects.all()
    
    context = {
        'wishes': wishes,
        'title': 'Gestion des souhaits de formation',
        'formations': all_formations,
        'current_formation_filter': formation_filter,
        'current_rpe_filter': rpe_filter,
        'current_date_from': date_from,
        'current_date_to': date_to,
    }
    
    return render(request, 'core/admin_training_wishes.html', context)

@staff_member_required
def assign_wish_to_session(request, wish_id):
    """Assigner un souhait de formation à une session."""
    wish = get_object_or_404(TrainingWish, id=wish_id)
    
    if request.method == 'POST':
        session_id = request.POST.get('session')
        if session_id:
            session = get_object_or_404(Session, id=session_id)
            
            # Créer un participant
            SessionParticipant.objects.create(
                session=session,
                user=wish.user,
                status='CONTACTED'  # Statut par défaut
            )
            
            # Créer une notification
            Notification.objects.create(
                user=wish.user,
                type='wish_assigned',
                message=f'Votre souhait pour la formation {wish.formation.name} a été assigné à une session',
                related_object_id=session.id,
                related_object_type='Session'
            )
            
            # Supprimer le souhait
            wish.delete()
            
            messages.success(request, 'Le souhait a été assigné à la session avec succès.')
            return redirect('core:admin_training_wishes')
    
    # Récupérer les sessions ouvertes pour cette formation
    sessions = Session.objects.filter(
        formation=wish.formation, 
        status__in=['NON_OUVERTE', 'DEMANDEE', 'OUVERTE']
    )
    
    context = {
        'wish': wish,
        'sessions': sessions,
        'title': 'Assigner un souhait à une session'
    }
    
    return render(request, 'core/assign_wish_to_session.html', context)


@login_required
@staff_member_required
def update_participant_status(request, participant_id):
    """Vue pour mettre à jour le statut d'un participant."""
    if request.method != 'POST':
        logger.error(f"Méthode non autorisée pour update_participant_status: {request.method}")
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        logger.info(f"Tentative de mise à jour du statut pour le participant {participant_id}")
        logger.info(f"Données POST reçues: {request.POST}")
        
        participant = SessionParticipant.objects.get(id=participant_id)
        new_status = request.POST.get('status')
        comment = request.POST.get('comment')
        
        logger.info(f"Nouveau statut: {new_status}")
        
        # Valider le statut
        valid_statuses = [
            SessionParticipant.STATUS_WISH,
            SessionParticipant.STATUS_CONTACTED,
            SessionParticipant.STATUS_REMINDED,
            SessionParticipant.STATUS_FILE_RECEIVED_EMAIL,
            SessionParticipant.STATUS_FILE_RECEIVED_PAPER,
            SessionParticipant.STATUS_ERROR
        ]
        
        if not new_status:
            logger.error("Aucun statut fourni")
            return JsonResponse({'error': 'Aucun statut fourni'}, status=400)
        
        if new_status not in valid_statuses:
            logger.error(f"Statut invalide: {new_status}")
            return JsonResponse({'error': 'Statut invalide'}, status=400)
        
        participant.status = new_status
        participant.save()
        
        # Ajouter un commentaire si fourni
        if comment:
            ParticipantComment.objects.create(
                participant=participant,
                author=request.user,
                content=comment
            )
        
        logger.info(f"Statut mis à jour avec succès pour le participant {participant_id}")
        
        return JsonResponse({
            'success': True,
            'status': participant.get_status_display(),
            'status_class': participant.get_status_badge_class(),
        })
    except SessionParticipant.DoesNotExist:
        logger.error(f"Participant non trouvé: {participant_id}")
        return JsonResponse({'error': 'Participant non trouvé'}, status=404)
    except Exception as e:
        logger.error(f"Erreur lors de la mise à jour du statut: {str(e)}")
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@staff_member_required
def update_participant_comments(request, participant_id):
    """Vue pour mettre à jour les commentaires d'un participant."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        participant = SessionParticipant.objects.get(id=participant_id)
        comment_content = request.POST.get('comments', '').strip()
        
        if not comment_content:
            return JsonResponse({'error': 'Le commentaire ne peut pas être vide'}, status=400)
            
        # Créer un nouveau commentaire
        ParticipantComment.objects.create(
            participant=participant,
            author=request.user,
            content=comment_content
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Commentaire ajouté avec succès'
        })
            
    except SessionParticipant.DoesNotExist:
        return JsonResponse({'error': 'Participant non trouvé'}, status=404)
    except Exception as e:
        print(f"DEBUG: Erreur lors de l'ajout du commentaire : {str(e)}")
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@staff_member_required
def get_session(request, session_id):
    """Vue pour récupérer les détails d'une session."""
    try:
        session = Session.objects.select_related('formation').prefetch_related(
            'dates', 
            'trainers'
        ).get(id=session_id)
        
        data = {
            'id': session.id,
            'formation': session.formation.id,
            'status': session.status,
            'iperia_opening_date': session.iperia_opening_date.strftime('%Y-%m-%d') if session.iperia_opening_date else None,
            'iperia_deadline': session.iperia_deadline.strftime('%Y-%m-%d') if session.iperia_deadline else None,
            'trainers': [trainer.id for trainer in session.trainers.all()],
            'dates': [{
                'date': date.date.strftime('%Y-%m-%d'),
                'location': date.location.id if date.location else None
            } for date in session.dates.all()]
        }
        return JsonResponse(data)
    except Session.DoesNotExist:
        return JsonResponse({'error': 'Session non trouvée'}, status=404)
    except Exception as e:
        logger.error(f"Erreur lors de la récupération de la session {session_id}: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@staff_member_required
def delete_session(request, session_id):
    """Vue pour supprimer une session."""
    try:
        session = Session.objects.get(id=session_id)
        session.delete()
        return JsonResponse({'success': True})
    except Session.DoesNotExist:
        return JsonResponse({'error': 'Session non trouvée'}, status=404)
    except Exception as e:
        logger.error(f"Erreur lors de la suppression de la session {session_id}: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@staff_member_required
def add_participant_from_wish(request, session_id):
    """Ajouter un participant depuis un voeu de formation."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        session = Session.objects.get(id=session_id)
        wish_id = request.POST.get('wish_id')
        training_wish = TrainingWish.objects.get(id=wish_id)
        user = training_wish.user
        
        # Créer le participant
        participant = SessionParticipant.objects.create(
            session=session,
            user=user,
            status='CONTACTED'  # Statut par défaut
        )
        
        # Créer une notification pour l'utilisateur
        Notification.objects.create(
            user=user,
            title='Inscription à une formation',
            message=f'Vous avez été inscrit à la formation "{session.formation.name}".',
            notification_type='SESSION_REGISTRATION'
        )
        
        # Mettre à jour le voeu de formation
        training_wish.session = session
        training_wish.save()
        
        return JsonResponse({
            'success': True,
            'participant': {
                'id': participant.id,
                'name': participant.user.get_full_name(),
                'email': participant.user.email,
                'status': participant.get_status_display(),
                'status_class': participant.get_status_badge_class(),
            }
        })
    except Session.DoesNotExist:
        return JsonResponse({'error': 'Session non trouvée'}, status=404)
    except TrainingWish.DoesNotExist:
        return JsonResponse({'error': 'Voeu de formation non trouvé'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def add_training_wish(request, formation_pk):
    """Ajouter un souhait de formation depuis la page de détail de formation."""
    formation = get_object_or_404(Formation, pk=formation_pk)
    
    if request.method == 'POST':
        try:
            # Créer le souhait de formation
            wish = TrainingWish.objects.create(
                user=request.user,
                formation=formation,
                notes=request.POST.get('message', '')
            )
            messages.success(request, 'Votre souhait de formation a été enregistré.')
        except IntegrityError:
            messages.error(request, 'Vous avez déjà exprimé un souhait pour cette formation.')
        
        return redirect('core:formation_detail', pk=formation_pk)

@login_required
def notifications_list(request):
    """Liste des notifications de l'utilisateur."""
    # Récupérer toutes les notifications non lues
    unread_notifications = request.user.notifications.filter(is_read=False)
    
    # Récupérer les notifications lues, limitées à 20
    read_notifications = request.user.notifications.filter(is_read=True).order_by('-created_at')[:20]
    
    # Combiner et trier les notifications
    notifications = list(unread_notifications) + list(read_notifications)
    
    # Marquer les notifications comme lues
    unread_notifications.update(is_read=True)
    
    context = {
        'notifications': notifications,
        'title': 'Mes notifications'
    }
    
    return render(request, 'core/notifications_list.html', context)


@staff_member_required
def admin_training_wishes(request):
    """Vue pour la gestion des souhaits de formation par les administrateurs."""
    # Récupérer tous les souhaits de formation
    wishes = TrainingWish.objects.select_related('user', 'formation').order_by('-created_at')
    
    # Filtres
    formation_filter = request.GET.get('formation', '')
    rpe_filter = request.GET.get('rpe', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Appliquer les filtres
    if formation_filter:
        wishes = wishes.filter(formation__name__icontains=formation_filter)
    
    if rpe_filter:
        wishes = wishes.filter(user__rpe_association__icontains=rpe_filter)
    
    if date_from:
        wishes = wishes.filter(created_at__date__gte=date_from)
    
    if date_to:
        wishes = wishes.filter(created_at__date__lte=date_to)
    
    # Récupérer toutes les formations pour le filtre
    all_formations = Formation.objects.all()
    
    context = {
        'wishes': wishes,
        'title': 'Gestion des souhaits de formation',
        'formations': all_formations,
        'current_formation_filter': formation_filter,
        'current_rpe_filter': rpe_filter,
        'current_date_from': date_from,
        'current_date_to': date_to,
    }
    
    return render(request, 'core/admin_training_wishes.html', context)

@staff_member_required
def assign_wish_to_session(request, wish_id):
    """Assigner un souhait de formation à une session."""
    wish = get_object_or_404(TrainingWish, id=wish_id)
    
    if request.method == 'POST':
        session_id = request.POST.get('session')
        if session_id:
            session = get_object_or_404(Session, id=session_id)
            
            # Créer un participant
            SessionParticipant.objects.create(
                session=session,
                user=wish.user,
                status='CONTACTED'  # Statut par défaut
            )
            
            # Créer une notification
            Notification.objects.create(
                user=wish.user,
                type='wish_assigned',
                message=f'Votre souhait pour la formation {wish.formation.name} a été assigné à une session',
                related_object_id=session.id,
                related_object_type='Session'
            )
            
            # Supprimer le souhait
            wish.delete()
            
            messages.success(request, 'Le souhait a été assigné à la session avec succès.')
            return redirect('core:admin_training_wishes')
    
    # Récupérer les sessions ouvertes pour cette formation
    sessions = Session.objects.filter(
        formation=wish.formation, 
        status__in=['NON_OUVERTE', 'DEMANDEE', 'OUVERTE']
    )
    
    context = {
        'wish': wish,
        'sessions': sessions,
        'title': 'Assigner un souhait à une session'
    }
    
    return render(request, 'core/assign_wish_to_session.html', context)


@login_required
@staff_member_required
def update_participant_status(request, participant_id):
    """Vue pour mettre à jour le statut d'un participant."""
    if request.method != 'POST':
        logger.error(f"Méthode non autorisée pour update_participant_status: {request.method}")
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        logger.info(f"Tentative de mise à jour du statut pour le participant {participant_id}")
        logger.info(f"Données POST reçues: {request.POST}")
        
        participant = SessionParticipant.objects.get(id=participant_id)
        new_status = request.POST.get('status')
        comment = request.POST.get('comment')
        
        logger.info(f"Nouveau statut: {new_status}")
        
        # Valider le statut
        valid_statuses = [
            SessionParticipant.STATUS_WISH,
            SessionParticipant.STATUS_CONTACTED,
            SessionParticipant.STATUS_REMINDED,
            SessionParticipant.STATUS_FILE_RECEIVED_EMAIL,
            SessionParticipant.STATUS_FILE_RECEIVED_PAPER,
            SessionParticipant.STATUS_ERROR
        ]
        
        if not new_status:
            logger.error("Aucun statut fourni")
            return JsonResponse({'error': 'Aucun statut fourni'}, status=400)
        
        if new_status not in valid_statuses:
            logger.error(f"Statut invalide: {new_status}")
            return JsonResponse({'error': 'Statut invalide'}, status=400)
        
        participant.status = new_status
        participant.save()
        
        # Ajouter un commentaire si fourni
        if comment:
            ParticipantComment.objects.create(
                participant=participant,
                author=request.user,
                content=comment
            )
        
        logger.info(f"Statut mis à jour avec succès pour le participant {participant_id}")
        
        return JsonResponse({
            'success': True,
            'status': participant.get_status_display(),
            'status_class': participant.get_status_badge_class(),
        })
    except SessionParticipant.DoesNotExist:
        logger.error(f"Participant non trouvé: {participant_id}")
        return JsonResponse({'error': 'Participant non trouvé'}, status=404)
    except Exception as e:
        logger.error(f"Erreur lors de la mise à jour du statut: {str(e)}")
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@staff_member_required
def update_participant_comments(request, participant_id):
    """Vue pour mettre à jour les commentaires d'un participant."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        participant = SessionParticipant.objects.get(id=participant_id)
        comment_content = request.POST.get('comments', '').strip()
        
        if not comment_content:
            return JsonResponse({'error': 'Le commentaire ne peut pas être vide'}, status=400)
            
        # Créer un nouveau commentaire
        ParticipantComment.objects.create(
            participant=participant,
            author=request.user,
            content=comment_content
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Commentaire ajouté avec succès'
        })
            
    except SessionParticipant.DoesNotExist:
        return JsonResponse({'error': 'Participant non trouvé'}, status=404)
    except Exception as e:
        print(f"DEBUG: Erreur lors de l'ajout du commentaire : {str(e)}")
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@staff_member_required
def get_session(request, session_id):
    """Vue pour récupérer les détails d'une session."""
    try:
        session = Session.objects.select_related('formation').prefetch_related(
            'dates', 
            'trainers'
        ).get(id=session_id)
        
        data = {
            'id': session.id,
            'formation': session.formation.id,
            'status': session.status,
            'iperia_opening_date': session.iperia_opening_date.strftime('%Y-%m-%d') if session.iperia_opening_date else None,
            'iperia_deadline': session.iperia_deadline.strftime('%Y-%m-%d') if session.iperia_deadline else None,
            'trainers': [trainer.id for trainer in session.trainers.all()],
            'dates': [{
                'date': date.date.strftime('%Y-%m-%d'),
                'location': date.location.id if date.location else None
            } for date in session.dates.all()]
        }
        return JsonResponse(data)
    except Session.DoesNotExist:
        return JsonResponse({'error': 'Session non trouvée'}, status=404)
    except Exception as e:
        logger.error(f"Erreur lors de la récupération de la session {session_id}: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@staff_member_required
def delete_session(request, session_id):
    """Vue pour supprimer une session."""
    try:
        session = Session.objects.get(id=session_id)
        session.delete()
        return JsonResponse({'success': True})
    except Session.DoesNotExist:
        return JsonResponse({'error': 'Session non trouvée'}, status=404)
    except Exception as e:
        logger.error(f"Erreur lors de la suppression de la session {session_id}: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@staff_member_required
def add_participant_from_wish(request, session_id):
    """Ajouter un participant depuis un voeu de formation."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        session = Session.objects.get(id=session_id)
        wish_id = request.POST.get('wish_id')
        training_wish = TrainingWish.objects.get(id=wish_id)
        user = training_wish.user
        
        # Créer le participant
        participant = SessionParticipant.objects.create(
            session=session,
            user=user,
            status='CONTACTED'  # Statut par défaut
        )
        
        # Créer une notification pour l'utilisateur
        Notification.objects.create(
            user=user,
            title='Inscription à une formation',
            message=f'Vous avez été inscrit à la formation "{session.formation.name}".',
            notification_type='SESSION_REGISTRATION'
        )
        
        # Mettre à jour le voeu de formation
        training_wish.session = session
        training_wish.save()
        
        return JsonResponse({
            'success': True,
            'participant': {
                'id': participant.id,
                'name': participant.user.get_full_name(),
                'email': participant.user.email,
                'status': participant.get_status_display(),
                'status_class': participant.get_status_badge_class(),
            }
        })
    except Session.DoesNotExist:
        return JsonResponse({'error': 'Session non trouvée'}, status=404)
    except TrainingWish.DoesNotExist:
        return JsonResponse({'error': 'Voeu de formation non trouvé'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def add_training_wish(request, formation_pk):
    """Ajouter un souhait de formation depuis la page de détail de formation."""
    formation = get_object_or_404(Formation, pk=formation_pk)
    
    if request.method == 'POST':
        try:
            # Créer le souhait de formation
            wish = TrainingWish.objects.create(
                user=request.user,
                formation=formation,
                notes=request.POST.get('message', '')
            )
            messages.success(request, 'Votre souhait de formation a été enregistré.')
        except IntegrityError:
            messages.error(request, 'Vous avez déjà exprimé un souhait pour cette formation.')
        
        return redirect('core:formation_detail', pk=formation_pk)

@login_required
def notifications_list(request):
    """Liste des notifications de l'utilisateur."""
    # Récupérer toutes les notifications non lues
    unread_notifications = request.user.notifications.filter(is_read=False)
    
    # Récupérer les notifications lues, limitées à 20
    read_notifications = request.user.notifications.filter(is_read=True).order_by('-created_at')[:20]
    
    # Combiner et trier les notifications
    notifications = list(unread_notifications) + list(read_notifications)
    
    # Marquer les notifications comme lues
    unread_notifications.update(is_read=True)
    
    context = {
        'notifications': notifications,
        'title': 'Mes notifications'
    }
    
    return render(request, 'core/notifications_list.html', context)


@staff_member_required
def admin_training_wishes(request):
    """Vue pour la gestion des souhaits de formation par les administrateurs."""
    # Récupérer tous les souhaits de formation
    wishes = TrainingWish.objects.select_related('user', 'formation').order_by('-created_at')
    
    # Filtres
    formation_filter = request.GET.get('formation', '')
    rpe_filter = request.GET.get('rpe', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Appliquer les filtres
    if formation_filter:
        wishes = wishes.filter(formation__name__icontains=formation_filter)
    
    if rpe_filter:
        wishes = wishes.filter(user__rpe_association__icontains=rpe_filter)
    
    if date_from:
        wishes = wishes.filter(created_at__date__gte=date_from)
    
    if date_to:
        wishes = wishes.filter(created_at__date__lte=date_to)
    
    # Récupérer toutes les formations pour le filtre
    all_formations = Formation.objects.all()
    
    context = {
        'wishes': wishes,
        'title': 'Gestion des souhaits de formation',
        'formations': all_formations,
        'current_formation_filter': formation_filter,
        'current_rpe_filter': rpe_filter,
        'current_date_from': date_from,
        'current_date_to': date_to,
    }
    
    return render(request, 'core/admin_training_wishes.html', context)

@staff_member_required
def assign_wish_to_session(request, wish_id):
    """Assigner un souhait de formation à une session."""
    wish = get_object_or_404(TrainingWish, id=wish_id)
    
    if request.method == 'POST':
        session_id = request.POST.get('session')
        if session_id:
            session = get_object_or_404(Session, id=session_id)
            
            # Créer un participant
            SessionParticipant.objects.create(
                session=session,
                user=wish.user,
                status='CONTACTED'  # Statut par défaut
            )
            
            # Créer une notification
            Notification.objects.create(
                user=wish.user,
                type='wish_assigned',
                message=f'Votre souhait pour la formation {wish.formation.name} a été assigné à une session',
                related_object_id=session.id,
                related_object_type='Session'
            )
            
            # Supprimer le souhait
            wish.delete()
            
            messages.success(request, 'Le souhait a été assigné à la session avec succès.')
            return redirect('core:admin_training_wishes')
    
    # Récupérer les sessions ouvertes pour cette formation
    sessions = Session.objects.filter(
        formation=wish.formation, 
        status__in=['NON_OUVERTE', 'DEMANDEE', 'OUVERTE']
    )
    
    context = {
        'wish': wish,
        'sessions': sessions,
        'title': 'Assigner un souhait à une session'
    }
    
    return render(request, 'core/assign_wish_to_session.html', context)


@login_required
@staff_member_required
def update_participant_status(request, participant_id):
    """Vue pour mettre à jour le statut d'un participant."""
    if request.method != 'POST':
        logger.error(f"Méthode non autorisée pour update_participant_status: {request.method}")
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        logger.info(f"Tentative de mise à jour du statut pour le participant {participant_id}")
        logger.info(f"Données POST reçues: {request.POST}")
        
        participant = SessionParticipant.objects.get(id=participant_id)
        new_status = request.POST.get('status')
        comment = request.POST.get('comment')
        
        logger.info(f"Nouveau statut: {new_status}")
        
        # Valider le statut
        valid_statuses = [
            SessionParticipant.STATUS_WISH,
            SessionParticipant.STATUS_CONTACTED,
            SessionParticipant.STATUS_REMINDED,
            SessionParticipant.STATUS_FILE_RECEIVED_EMAIL,
            SessionParticipant.STATUS_FILE_RECEIVED_PAPER,
            SessionParticipant.STATUS_ERROR
        ]
        
        if not new_status:
            logger.error("Aucun statut fourni")
            return JsonResponse({'error': 'Aucun statut fourni'}, status=400)
        
        if new_status not in valid_statuses:
            logger.error(f"Statut invalide: {new_status}")
            return JsonResponse({'error': 'Statut invalide'}, status=400)
        
        participant.status = new_status
        participant.save()
        
        # Ajouter un commentaire si fourni
        if comment:
            ParticipantComment.objects.create(
                participant=participant,
                author=request.user,
                content=comment
            )
        
        logger.info(f"Statut mis à jour avec succès pour le participant {participant_id}")
        
        return JsonResponse({
            'success': True,
            'status': participant.get_status_display(),
            'status_class': participant.get_status_badge_class(),
        })
    except SessionParticipant.DoesNotExist:
        logger.error(f"Participant non trouvé: {participant_id}")
        return JsonResponse({'error': 'Participant non trouvé'}, status=404)
    except Exception as e:
        logger.error(f"Erreur lors de la mise à jour du statut: {str(e)}")
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@staff_member_required
def update_participant_comments(request, participant_id):
    """Vue pour mettre à jour les commentaires d'un participant."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        participant = SessionParticipant.objects.get(id=participant_id)
        comment_content = request.POST.get('comments', '').strip()
        
        if not comment_content:
            return JsonResponse({'error': 'Le commentaire ne peut pas être vide'}, status=400)
            
        # Créer un nouveau commentaire
        ParticipantComment.objects.create(
            participant=participant,
            author=request.user,
            content=comment_content
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Commentaire ajouté avec succès'
        })
            
    except SessionParticipant.DoesNotExist:
        return JsonResponse({'error': 'Participant non trouvé'}, status=404)
    except Exception as e:
        print(f"DEBUG: Erreur lors de l'ajout du commentaire : {str(e)}")
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@staff_member_required
def get_session(request, session_id):
    """Vue pour récupérer les détails d'une session."""
    try:
        session = Session.objects.select_related('formation').prefetch_related(
            'dates', 
            'trainers'
        ).get(id=session_id)
        
        data = {
            'id': session.id,
            'formation': session.formation.id,
            'status': session.status,
            'iperia_opening_date': session.iperia_opening_date.strftime('%Y-%m-%d') if session.iperia_opening_date else None,
            'iperia_deadline': session.iperia_deadline.strftime('%Y-%m-%d') if session.iperia_deadline else None,
            'trainers': [trainer.id for trainer in session.trainers.all()],
            'dates': [{
                'date': date.date.strftime('%Y-%m-%d'),
                'location': date.location.id if date.location else None
            } for date in session.dates.all()]
        }
        return JsonResponse(data)
    except Session.DoesNotExist:
        return JsonResponse({'error': 'Session non trouvée'}, status=404)
    except Exception as e:
        logger.error(f"Erreur lors de la récupération de la session {session_id}: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@staff_member_required
def delete_session(request, session_id):
    """Vue pour supprimer une session."""
    try:
        session = Session.objects.get(id=session_id)
        session.delete()
        return JsonResponse({'success': True})
    except Session.DoesNotExist:
        return JsonResponse({'error': 'Session non trouvée'}, status=404)
    except Exception as e:
        logger.error(f"Erreur lors de la suppression de la session {session_id}: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@staff_member_required
def add_participant_from_wish(request, session_id):
    """Ajouter un participant depuis un voeu de formation."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        session = Session.objects.get(id=session_id)
        wish_id = request.POST.get('wish_id')
        training_wish = TrainingWish.objects.get(id=wish_id)
        user = training_wish.user
        
        # Créer le participant
        participant = SessionParticipant.objects.create(
            session=session,
            user=user,
            status='CONTACTED'  # Statut par défaut
        )
        
        # Créer une notification pour l'utilisateur
        Notification.objects.create(
            user=user,
            title='Inscription à une formation',
            message=f'Vous avez été inscrit à la formation "{session.formation.name}".',
            notification_type='SESSION_REGISTRATION'
        )
        
        # Mettre à jour le voeu de formation
        training_wish.session = session
        training_wish.save()
        
        return JsonResponse({
            'success': True,
            'participant': {
                'id': participant.id,
                'name': participant.user.get_full_name(),
                'email': participant.user.email,
                'status': participant.get_status_display(),
                'status_class': participant.get_status_badge_class(),
            }
        })
    except Session.DoesNotExist:
        return JsonResponse({'error': 'Session non trouvée'}, status=404)
    except TrainingWish.DoesNotExist:
        return JsonResponse({'error': 'Voeu de formation non trouvé'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def add_training_wish(request, formation_pk):
    """Ajouter un souhait de formation depuis la page de détail de formation."""
    formation = get_object_or_404(Formation, pk=formation_pk)
    
    if request.method == 'POST':
        try:
            # Créer le souhait de formation
            wish = TrainingWish.objects.create(
                user=request.user,
                formation=formation,
                notes=request.POST.get('message', '')
            )
            messages.success(request, 'Votre souhait de formation a été enregistré.')
        except IntegrityError:
            messages.error(request, 'Vous avez déjà exprimé un souhait pour cette formation.')
        
        return redirect('core:formation_detail', pk=formation_pk)

@login_required
def notifications_list(request):
    """Liste des notifications de l'utilisateur."""
    # Récupérer toutes les notifications non lues
    unread_notifications = request.user.notifications.filter(is_read=False)
    
    # Récupérer les notifications lues, limitées à 20
    read_notifications = request.user.notifications.filter(is_read=True).order_by('-created_at')[:20]
    
    # Combiner et trier les notifications
    notifications = list(unread_notifications) + list(read_notifications)
    
    # Marquer les notifications comme lues
    unread_notifications.update(is_read=True)
    
    context = {
        'notifications': notifications,
        'title': 'Mes notifications'
    }
    
    return render(request, 'core/notifications_list.html', context)


@staff_member_required
def admin_training_wishes(request):
    """Vue pour la gestion des souhaits de formation par les administrateurs."""
    # Récupérer tous les souhaits de formation
    wishes = TrainingWish.objects.select_related('user', 'formation').order_by('-created_at')
    
    # Filtres
    formation_filter = request.GET.get('formation', '')
    rpe_filter = request.GET.get('rpe', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Appliquer les filtres
    if formation_filter:
        wishes = wishes.filter(formation__name__icontains=formation_filter)
    
    if rpe_filter:
        wishes = wishes.filter(user__rpe_association__icontains=rpe_filter)
    
    if date_from:
        wishes = wishes.filter(created_at__date__gte=date_from)
    
    if date_to:
        wishes = wishes.filter(created_at__date__lte=date_to)
    
    # Récupérer toutes les formations pour le filtre
    all_formations = Formation.objects.all()
    
    context = {
        'wishes': wishes,
        'title': 'Gestion des souhaits de formation',
        'formations': all_formations,
        'current_formation_filter': formation_filter,
        'current_rpe_filter': rpe_filter,
        'current_date_from': date_from,
        'current_date_to': date_to,
    }
    
    return render(request, 'core/admin_training_wishes.html', context)

@staff_member_required
def assign_wish_to_session(request, wish_id):
    """Assigner un souhait de formation à une session."""
    wish = get_object_or_404(TrainingWish, id=wish_id)
    
    if request.method == 'POST':
        session_id = request.POST.get('session')
        if session_id:
            session = get_object_or_404(Session, id=session_id)
            
            # Créer un participant
            SessionParticipant.objects.create(
                session=session,
                user=wish.user,
                status='CONTACTED'  # Statut par défaut
            )
            
            # Créer une notification
            Notification.objects.create(
                user=wish.user,
                type='wish_assigned',
                message=f'Votre souhait pour la formation {wish.formation.name} a été assigné à une session',
                related_object_id=session.id,
                related_object_type='Session'
            )
            
            # Supprimer le souhait
            wish.delete()
            
            messages.success(request, 'Le souhait a été assigné à la session avec succès.')
            return redirect('core:admin_training_wishes')
    
    # Récupérer les sessions ouvertes pour cette formation
    sessions = Session.objects.filter(
        formation=wish.formation, 
        status__in=['NON_OUVERTE', 'DEMANDEE', 'OUVERTE']
    )
    
    context = {
        'wish': wish,
        'sessions': sessions,
        'title': 'Assigner un souhait à une session'
    }
    
    return render(request, 'core/assign_wish_to_session.html', context)


@login_required
@staff_member_required
def update_participant_status(request, participant_id):
    """Vue pour mettre à jour le statut d'un participant."""
    if request.method != 'POST':
        logger.error(f"Méthode non autorisée pour update_participant_status: {request.method}")
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        logger.info(f"Tentative de mise à jour du statut pour le participant {participant_id}")
        logger.info(f"Données POST reçues: {request.POST}")
        
        participant = SessionParticipant.objects.get(id=participant_id)
        new_status = request.POST.get('status')
        comment = request.POST.get('comment')
        
        logger.info(f"Nouveau statut: {new_status}")
        
        # Valider le statut
        valid_statuses = [
            SessionParticipant.STATUS_WISH,
            SessionParticipant.STATUS_CONTACTED,
            SessionParticipant.STATUS_REMINDED,
            SessionParticipant.STATUS_FILE_RECEIVED_EMAIL,
            SessionParticipant.STATUS_FILE_RECEIVED_PAPER,
            SessionParticipant.STATUS_ERROR
        ]
        
        if not new_status:
            logger.error("Aucun statut fourni")
            return JsonResponse({'error': 'Aucun statut fourni'}, status=400)
        
        if new_status not in valid_statuses:
            logger.error(f"Statut invalide: {new_status}")
            return JsonResponse({'error': 'Statut invalide'}, status=400)
        
        participant.status = new_status
        participant.save()
        
        # Ajouter un commentaire si fourni
        if comment:
            ParticipantComment.objects.create(
                participant=participant,
                author=request.user,
                content=comment
            )
        
        logger.info(f"Statut mis à jour avec succès pour le participant {participant_id}")
        
        return JsonResponse({
            'success': True,
            'status': participant.get_status_display(),
            'status_class': participant.get_status_badge_class(),
        })
    except SessionParticipant.DoesNotExist:
        logger.error(f"Participant non trouvé: {participant_id}")
        return JsonResponse({'error': 'Participant non trouvé'}, status=404)
    except Exception as e:
        logger.error(f"Erreur lors de la mise à jour du statut: {str(e)}")
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@staff_member_required
def update_participant_comments(request, participant_id):
    """Vue pour mettre à jour les commentaires d'un participant."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        participant = SessionParticipant.objects.get(id=participant_id)
        comment_content = request.POST.get('comments', '').strip()
        
        if not comment_content:
            return JsonResponse({'error': 'Le commentaire ne peut pas être vide'}, status=400)
            
        # Créer un nouveau commentaire
        ParticipantComment.objects.create(
            participant=participant,
            author=request.user,
            content=comment_content
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Commentaire ajouté avec succès'
        })
            
    except SessionParticipant.DoesNotExist:
        return JsonResponse({'error': 'Participant non trouvé'}, status=404)
    except Exception as e:
        print(f"DEBUG: Erreur lors de l'ajout du commentaire : {str(e)}")
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@staff_member_required
def get_session(request, session_id):
    """Vue pour récupérer les détails d'une session."""
    try:
        session = Session.objects.select_related('formation').prefetch_related(
            'dates', 
            'trainers'
        ).get(id=session_id)
        
        data = {
            'id': session.id,
            'formation': session.formation.id,
            'status': session.status,
            'iperia_opening_date': session.iperia_opening_date.strftime('%Y-%m-%d') if session.iperia_opening_date else None,
            'iperia_deadline': session.iperia_deadline.strftime('%Y-%m-%d') if session.iperia_deadline else None,
            'trainers': [trainer.id for trainer in session.trainers.all()],
            'dates': [{
                'date': date.date.strftime('%Y-%m-%d'),
                'location': date.location.id if date.location else None
            } for date in session.dates.all()]
        }
        return JsonResponse(data)
    except Session.DoesNotExist:
        return JsonResponse({'error': 'Session non trouvée'}, status=404)
    except Exception as e:
        logger.error(f"Erreur lors de la récupération de la session {session_id}: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@staff_member_required
def delete_session(request, session_id):
    """Vue pour supprimer une session."""
    try:
        session = Session.objects.get(id=session_id)
        session.delete()
        return JsonResponse({'success': True})
    except Session.DoesNotExist:
        return JsonResponse({'error': 'Session non trouvée'}, status=404)
    except Exception as e:
        logger.error(f"Erreur lors de la suppression de la session {session_id}: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@staff_member_required
def add_participant_from_wish(request, session_id):
    """Ajouter un participant depuis un voeu de formation."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        session = Session.objects.get(id=session_id)
        wish_id = request.POST.get('wish_id')
        training_wish = TrainingWish.objects.get(id=wish_id)
        user = training_wish.user
        
        # Créer le participant
        participant = SessionParticipant.objects.create(
            session=session,
            user=user,
            status='CONTACTED'  # Statut par défaut
        )
        
        # Créer une notification pour l'utilisateur
        Notification.objects.create(
            user=user,
            title='Inscription à une formation',
            message=f'Vous avez été inscrit à la formation "{session.formation.name}".',
            notification_type='SESSION_REGISTRATION'
        )
        
        # Mettre à jour le voeu de formation
        training_wish.session = session
        training_wish.save()
        
        return JsonResponse({
            'success': True,
            'participant': {
                'id': participant.id,
                'name': participant.user.get_full_name(),
                'email': participant.user.email,
                'status': participant.get_status_display(),
                'status_class': participant.get_status_badge_class(),
            }
        })
    except Session.DoesNotExist:
        return JsonResponse({'error': 'Session non trouvée'}, status=404)
    except TrainingWish.DoesNotExist:
        return JsonResponse({'error': 'Voeu de formation non trouvé'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def add_training_wish(request, formation_pk):
    """Ajouter un souhait de formation depuis la page de détail de formation."""
    formation = get_object_or_404(Formation, pk=formation_pk)
    
    if request.method == 'POST':
        try:
            # Créer le souhait de formation
            wish = TrainingWish.objects.create(
                user=request.user,
                formation=formation,
                notes=request.POST.get('message', '')
            )
            messages.success(request, 'Votre souhait de formation a été enregistré.')
        except IntegrityError:
            messages.error(request, 'Vous avez déjà exprimé un souhait pour cette formation.')
        
        return redirect('core:formation_detail', pk=formation_pk)

@login_required
def notifications_list(request):
    """Liste des notifications de l'utilisateur."""
    # Récupérer toutes les notifications non lues
    unread_notifications = request.user.notifications.filter(is_read=False)
    
    # Récupérer les notifications lues, limitées à 20
    read_notifications = request.user.notifications.filter(is_read=True).order_by('-created_at')[:20]
    
    # Combiner et trier les notifications
    notifications = list(unread_notifications) + list(read_notifications)
    
    # Marquer les notifications comme lues
    unread_notifications.update(is_read=True)
    
    context = {
        'notifications': notifications,
        'title': 'Mes notifications'
    }
    
    return render(request, 'core/notifications_list.html', context)


@staff_member_required
def admin_training_wishes(request):
    """Vue pour la gestion des souhaits de formation par les administrateurs."""
    # Récupérer tous les souhaits de formation
    wishes = TrainingWish.objects.select_related('user', 'formation').order_by('-created_at')
    
    # Filtres
    formation_filter = request.GET.get('formation', '')
    rpe_filter = request.GET.get('rpe', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Appliquer les filtres
    if formation_filter:
        wishes = wishes.filter(formation__name__icontains=formation_filter)
    
    if rpe_filter:
        wishes = wishes.filter(user__rpe_association__icontains=rpe_filter)
    
    if date_from:
        wishes = wishes.filter(created_at__date__gte=date_from)
    
    if date_to:
        wishes = wishes.filter(created_at__date__lte=date_to)
    
    # Récupérer toutes les formations pour le filtre
    all_formations = Formation.objects.all()
    
    context = {
        'wishes': wishes,
        'title': 'Gestion des souhaits de formation',
        'formations': all_formations,
        'current_formation_filter': formation_filter,
        'current_rpe_filter': rpe_filter,
        'current_date_from': date_from,
        'current_date_to': date_to,
    }
    
    return render(request, 'core/admin_training_wishes.html', context)

@staff_member_required
def assign_wish_to_session(request, wish_id):
    """Assigner un souhait de formation à une session."""
    wish = get_object_or_404(TrainingWish, id=wish_id)
    
    if request.method == 'POST':
        session_id = request.POST.get('session')
        if session_id:
            session = get_object_or_404(Session, id=session_id)
            
            # Créer un participant
            SessionParticipant.objects.create(
                session=session,
                user=wish.user,
                status='CONTACTED'  # Statut par défaut
            )
            
            # Créer une notification
            Notification.objects.create(
                user=wish.user,
                type='wish_assigned',
                message=f'Votre souhait pour la formation {wish.formation.name} a été assigné à une session',
                related_object_id=session.id,
                related_object_type='Session'
            )
            
            # Supprimer le souhait
            wish.delete()
            
            messages.success(request, 'Le souhait a été assigné à la session avec succès.')
            return redirect('core:admin_training_wishes')
    
    # Récupérer les sessions ouvertes pour cette formation
    sessions = Session.objects.filter(
        formation=wish.formation, 
        status__in=['NON_OUVERTE', 'DEMANDEE', 'OUVERTE']
    )
    
    context = {
        'wish': wish,
        'sessions': sessions,
        'title': 'Assigner un souhait à une session'
    }
    
    return render(request, 'core/assign_wish_to_session.html', context)


@login_required
@staff_member_required
def update_participant_status(request, participant_id):
    """Vue pour mettre à jour le statut d'un participant."""
    if request.method != 'POST':
        logger.error(f"Méthode non autorisée pour update_participant_status: {request.method}")
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        logger.info(f"Tentative de mise à jour du statut pour le participant {participant_id}")
        logger.info(f"Données POST reçues: {request.POST}")
        
        participant = SessionParticipant.objects.get(id=participant_id)
        new_status = request.POST.get('status')
        comment = request.POST.get('comment')
        
        logger.info(f"Nouveau statut: {new_status}")
        
        # Valider le statut
        valid_statuses = [
            SessionParticipant.STATUS_WISH,
            SessionParticipant.STATUS_CONTACTED,
            SessionParticipant.STATUS_REMINDED,
            SessionParticipant.STATUS_FILE_RECEIVED_EMAIL,
            SessionParticipant.STATUS_FILE_RECEIVED_PAPER,
            SessionParticipant.STATUS_ERROR
        ]
        
        if not new_status:
            logger.error("Aucun statut fourni")
            return JsonResponse({'error': 'Aucun statut fourni'}, status=400)
        
        if new_status not in valid_statuses:
            logger.error(f"Statut invalide: {new_status}")
            return JsonResponse({'error': 'Statut invalide'}, status=400)
        
        participant.status = new_status
        participant.save()
        
        # Ajouter un commentaire si fourni
        if comment:
            ParticipantComment.objects.create(
                participant=participant,
                author=request.user,
                content=comment
            )
        
        logger.info(f"Statut mis à jour avec succès pour le participant {participant_id}")
        
        return JsonResponse({
            'success': True,
            'status': participant.get_status_display(),
            'status_class': participant.get_status_badge_class(),
        })
    except SessionParticipant.DoesNotExist:
        logger.error(f"Participant non trouvé: {participant_id}")
        return JsonResponse({'error': 'Participant non trouvé'}, status=404)
    except Exception as e:
        logger.error(f"Erreur lors de la mise à jour du statut: {str(e)}")
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@staff_member_required
def update_participant_comments(request, participant_id):
    """Vue pour mettre à jour les commentaires d'un participant."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        participant = SessionParticipant.objects.get(id=participant_id)
        comment_content = request.POST.get('comments', '').strip()
        
        if not comment_content:
            return JsonResponse({'error': 'Le commentaire ne peut pas être vide'}, status=400)
            
        # Créer un nouveau commentaire
        ParticipantComment.objects.create(
            participant=participant,
            author=request.user,
            content=comment_content
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Commentaire ajouté avec succès'
        })
            
    except SessionParticipant.DoesNotExist:
        return JsonResponse({'error': 'Participant non trouvé'}, status=404)
    except Exception as e:
        print(f"DEBUG: Erreur lors de l'ajout du commentaire : {str(e)}")
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@staff_member_required
def get_session(request, session_id):
    """Vue pour récupérer les détails d'une session."""
    try:
        session = Session.objects.select_related('formation').prefetch_related(
            'dates', 
            'trainers'
        ).get(id=session_id)
        
        data = {
            'id': session.id,
            'formation': session.formation.id,
            'status': session.status,
            'iperia_opening_date': session.iperia_opening_date.strftime('%Y-%m-%d') if session.iperia_opening_date else None,
            'iperia_deadline': session.iperia_deadline.strftime('%Y-%m-%d') if session.iperia_deadline else None,
            'trainers': [trainer.id for trainer in session.trainers.all()],
            'dates': [{
                'date': date.date.strftime('%Y-%m-%d'),
                'location': date.location.id if date.location else None
            } for date in session.dates.all()]
        }
        return JsonResponse(data)
    except Session.DoesNotExist:
        return JsonResponse({'error': 'Session non trouvée'}, status=404)
    except Exception as e:
        logger.error(f"Erreur lors de la récupération de la session {session_id}: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@staff_member_required
def delete_session(request, session_id):
    """Vue pour supprimer une session."""
    try:
        session = Session.objects.get(id=session_id)
        session.delete()
        return JsonResponse({'success': True})
    except Session.DoesNotExist:
        return JsonResponse({'error': 'Session non trouvée'}, status=404)
    except Exception as e:
        logger.error(f"Erreur lors de la suppression de la session {session_id}: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@staff_member_required
def add_participant_from_wish(request, session_id):
    """Ajouter un participant depuis un voeu de formation."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        session = Session.objects.get(id=session_id)
        wish_id = request.POST.get('wish_id')
        training_wish = TrainingWish.objects.get(id=wish_id)
        user = training_wish.user
        
        # Créer le participant
        participant = SessionParticipant.objects.create(
            session=session,
            user=user,
            status='CONTACTED'  # Statut par défaut
        )
        
        # Créer une notification pour l'utilisateur
        Notification.objects.create(
            user=user,
            title='Inscription à une formation',
            message=f'Vous avez été inscrit à la formation "{session.formation.name}".',
            notification_type='SESSION_REGISTRATION'
        )
        
        # Mettre à jour le voeu de formation
        training_wish.session = session
        training_wish.save()
        
        return JsonResponse({
            'success': True,
            'participant': {
                'id': participant.id,
                'name': participant.user.get_full_name(),
                'email': participant.user.email,
                'status': participant.get_status_display(),
                'status_class': participant.get_status_badge_class(),
            }
        })
    except Session.DoesNotExist:
        return JsonResponse({'error': 'Session non trouvée'}, status=404)
    except TrainingWish.DoesNotExist:
        return JsonResponse({'error': 'Voeu de formation non trouvé'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def add_training_wish(request, formation_pk):
    """Ajouter un souhait de formation depuis la page de détail de formation."""
    formation = get_object_or_404(Formation, pk=formation_pk)
    
    if request.method == 'POST':
        try:
            # Créer le souhait de formation
            wish = TrainingWish.objects.create(
                user=request.user,
                formation=formation,
                notes=request.POST.get('message', '')
            )
            messages.success(request, 'Votre souhait de formation a été enregistré.')
        except IntegrityError:
            messages.error(request, 'Vous avez déjà exprimé un souhait pour cette formation.')
        
        return redirect('core:formation_detail', pk=formation_pk)

@login_required
def notifications_list(request):
    """Liste des notifications de l'utilisateur."""
    # Récupérer toutes les notifications non lues
    unread_notifications = request.user.notifications.filter(is_read=False)
    
    # Récupérer les notifications lues, limitées à 20
    read_notifications = request.user.notifications.filter(is_read=True).order_by('-created_at')[:20]
    
    # Combiner et trier les notifications
    notifications = list(unread_notifications) + list(read_notifications)
    
    # Marquer les notifications comme lues
    unread_notifications.update(is_read=True)
    
    context = {
        'notifications': notifications,
        'title': 'Mes notifications'
    }
    
    return render(request, 'core/notifications_list.html', context)


@staff_member_required
def admin_training_wishes(request):
    """Vue pour la gestion des souhaits de formation par les administrateurs."""
    # Récupérer tous les souhaits de formation
    wishes = TrainingWish.objects.select_related('user', 'formation').order_by('-created_at')
    
    # Filtres
    formation_filter = request.GET.get('formation', '')
    rpe_filter = request.GET.get('rpe', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Appliquer les filtres
    if formation_filter:
        wishes = wishes.filter(formation__name__icontains=formation_filter)
    
    if rpe_filter:
        wishes = wishes.filter(user__rpe_association__icontains=rpe_filter)
    
    if date_from:
        wishes = wishes.filter(created_at__date__gte=date_from)
    
    if date_to:
        wishes = wishes.filter(created_at__date__lte=date_to)
    
    # Récupérer toutes les formations pour le filtre
    all_formations = Formation.objects.all()
    
    context = {
        'wishes': wishes,
        'title': 'Gestion des souhaits de formation',
        'formations': all_formations,
        'current_formation_filter': formation_filter,
        'current_rpe_filter': rpe_filter,
        'current_date_from': date_from,
        'current_date_to': date_to,
    }
    
    return render(request, 'core/admin_training_wishes.html', context)

@staff_member_required
def assign_wish_to_session(request, wish_id):
    """Assigner un souhait de formation à une session."""
    wish = get_object_or_404(TrainingWish, id=wish_id)
    
    if request.method == 'POST':
        session_id = request.POST.get('session')
        if session_id:
            session = get_object_or_404(Session, id=session_id)
            
            # Créer un participant
            SessionParticipant.objects.create(
                session=session,
                user=wish.user,
                status='CONTACTED'  # Statut par défaut
            )
            
            # Créer une notification
            Notification.objects.create(
                user=wish.user,
                type='wish_assigned',
                message=f'Votre souhait pour la formation {wish.formation.name} a été assigné à une session',
                related_object_id=session.id,
                related_object_type='Session'
            )
            
            # Supprimer le souhait
            wish.delete()
            
            messages.success(request, 'Le souhait a été assigné à la session avec succès.')
            return redirect('core:admin_training_wishes')
    
    # Récupérer les sessions ouvertes pour cette formation
    sessions = Session.objects.filter(
        formation=wish.formation, 
        status__in=['NON_OUVERTE', 'DEMANDEE', 'OUVERTE']
    )
    
    context = {
        'wish': wish,
        'sessions': sessions,
        'title': 'Assigner un souhait à une session'
    }
    
    return render(request, 'core/assign_wish_to_session.html', context)


@login_required
@staff_member_required
def update_participant_status(request, participant_id):
    """Vue pour mettre à jour le statut d'un participant."""
    if request.method != 'POST':
        logger.error(f"Méthode non autorisée pour update_participant_status: {request.method}")
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        logger.info(f"Tentative de mise à jour du statut pour le participant {participant_id}")
        logger.info(f"Données POST reçues: {request.POST}")
        
        participant = SessionParticipant.objects.get(id=participant_id)
        new_status = request.POST.get('status')
        comment = request.POST.get('comment')
        
        logger.info(f"Nouveau statut: {new_status}")
        
        # Valider le statut
        valid_statuses = [
            SessionParticipant.STATUS_WISH,
            SessionParticipant.STATUS_CONTACTED,
            SessionParticipant.STATUS_REMINDED,
            SessionParticipant.STATUS_FILE_RECEIVED_EMAIL,
            SessionParticipant.STATUS_FILE_RECEIVED_PAPER,
            SessionParticipant.STATUS_ERROR
        ]
        
        if not new_status:
            logger.error("Aucun statut fourni")
            return JsonResponse({'error': 'Aucun statut fourni'}, status=400)
        
        if new_status not in valid_statuses:
            logger.error(f"Statut invalide: {new_status}")
            return JsonResponse({'error': 'Statut invalide'}, status=400)
        
        participant.status = new_status
        participant.save()
        
        # Ajouter un commentaire si fourni
        if comment:
            ParticipantComment.objects.create(
                participant=participant,
                author=request.user,
                content=comment
            )
        
        logger.info(f"Statut mis à jour avec succès pour le participant {participant_id}")
        
        return JsonResponse({
            'success': True,
            'status': participant.get_status_display(),
            'status_class': participant.get_status_badge_class(),
        })
    except SessionParticipant.DoesNotExist:
        logger.error(f"Participant non trouvé: {participant_id}")
        return JsonResponse({'error': 'Participant non trouvé'}, status=404)
    except Exception as e:
        logger.error(f"Erreur lors de la mise à jour du statut: {str(e)}")
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@staff_member_required
def update_participant_comments(request, participant_id):
    """Vue pour mettre à jour les commentaires d'un participant."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        participant = SessionParticipant.objects.get(id=participant_id)
        comment_content = request.POST.get('comments', '').strip()
        
        if not comment_content:
            return JsonResponse({'error': 'Le commentaire ne peut pas être vide'}, status=400)
            
        # Créer un nouveau commentaire
        ParticipantComment.objects.create(
            participant=participant,
            author=request.user,
            content=comment_content
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Commentaire ajouté avec succès'
        })
            
    except SessionParticipant.DoesNotExist:
        return JsonResponse({'error': 'Participant non trouvé'}, status=404)
    except Exception as e:
        print(f"DEBUG: Erreur lors de l'ajout du commentaire : {str(e)}")
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@staff_member_required
def get_session(request, session_id):
    """Vue pour récupérer les détails d'une session."""
    try:
        session = Session.objects.select_related('formation').prefetch_related(
            'dates', 
            'trainers'
        ).get(id=session_id)
        
        data = {
            'id': session.id,
            'formation': session.formation.id,
            'status': session.status,
            'iperia_opening_date': session.iperia_opening_date.strftime('%Y-%m-%d') if session.iperia_opening_date else None,
            'iperia_deadline': session.iperia_deadline.strftime('%Y-%m-%d') if session.iperia_deadline else None,
            'trainers': [trainer.id for trainer in session.trainers.all()],
            'dates': [{
                'date': date.date.strftime('%Y-%m-%d'),
                'location': date.location.id if date.location else None
            } for date in session.dates.all()]
        }
        return JsonResponse(data)
    except Session.DoesNotExist:
        return JsonResponse({'error': 'Session non trouvée'}, status=404)
    except Exception as e:
        logger.error(f"Erreur lors de la récupération de la session {session_id}: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@staff_member_required
def delete_session(request, session_id):
    """Vue pour supprimer une session."""
    try:
        session = Session.objects.get(id=session_id)
        session.delete()
        return JsonResponse({'success': True})
    except Session.DoesNotExist:
        return JsonResponse({'error': 'Session non trouvée'}, status=404)
    except Exception as e:
        logger.error(f"Erreur lors de la suppression de la session {session_id}: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@staff_member_required
def add_participant_from_wish(request, session_id):
    """Ajouter un participant depuis un voeu de formation."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        session = Session.objects.get(id=session_id)
        wish_id = request.POST.get('wish_id')
        training_wish = TrainingWish.objects.get(id=wish_id)
        user = training_wish.user
        
        # Créer le participant
        participant = SessionParticipant.objects.create(
            session=session,
            user=user,
            status='CONTACTED'  # Statut par défaut
        )
        
        # Créer une notification pour l'utilisateur
        Notification.objects.create(
            user=user,
            title='Inscription à une formation',
            message=f'Vous avez été inscrit à la formation "{session.formation.name}".',
            notification_type='SESSION_REGISTRATION'
        )
        
        # Mettre à jour le voeu de formation
        training_wish.session = session
        training_wish.save()
        
        return JsonResponse({
            'success': True,
            'participant': {
                'id': participant.id,
                'name': participant.user.get_full_name(),
                'email': participant.user.email,
                'status': participant.get_status_display(),
                'status_class': participant.get_status_badge_class(),
            }
        })
    except Session.DoesNotExist:
        return JsonResponse({'error': 'Session non trouvée'}, status=404)
    except TrainingWish.DoesNotExist:
        return JsonResponse({'error': 'Voeu de formation non trouvé'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def add_training_wish(request, formation_pk):
    """Ajouter un souhait de formation depuis la page de détail de formation."""
    formation = get_object_or_404(Formation, pk=formation_pk)
    
    if request.method == 'POST':
        try:
            # Créer le souhait de formation
            wish = TrainingWish.objects.create(
                user=request.user,
                formation=formation,
                notes=request.POST.get('message', '')
            )
            messages.success(request, 'Votre souhait de formation a été enregistré.')
        except IntegrityError:
            messages.error(request, 'Vous avez déjà exprimé un souhait pour cette formation.')
        
        return redirect('core:formation_detail', pk=formation_pk)

@login_required
def notifications_list(request):
    """Liste des notifications de l'utilisateur."""
    # Récupérer toutes les notifications non lues
    unread_notifications = request.user.notifications.filter(is_read=False)
    
    # Récupérer les notifications lues, limitées à 20
    read_notifications = request.user.notifications.filter(is_read=True).order_by('-created_at')[:20]
    
    # Combiner et trier les notifications
    notifications = list(unread_notifications) + list(read_notifications)
    
    # Marquer les notifications comme lues
    unread_notifications.update(is_read=True)
    
    context = {
        'notifications': notifications,
        'title': 'Mes notifications'
    }
    
    return render(request, 'core/notifications_list.html', context)

def sessions_calendar(request):
    selected_formation = request.GET.get('formation')
    
    sessions = Session.objects.prefetch_related(
        'trainers',
        Prefetch('dates')
    ).select_related('formation')

    if selected_formation:
        sessions = sessions.filter(formation_id=selected_formation)

    formations = Formation.objects.all()

    context = {
        'title': 'Calendrier des sessions',
        'sessions': sessions,
        'formations': formations,
        'selected_formation': selected_formation,
    }

    return render(request, 'core/sessions_calendar.html', context)
@login_required
def get_participant_comments(request, participant_id):
    """
    Vue pour récupérer les commentaires d'un participant.
    """
    participant = get_object_or_404(SessionParticipant, id=participant_id)
    comments = ParticipantComment.objects.filter(participant=participant).order_by('-created_at')
    
    comments_data = [{
        'id': comment.id,
        'content': comment.content,  # ✅ Correction ici
        'created_at': comment.created_at.strftime('%d/%m/%Y %H:%M'),
        'author': comment.author.get_full_name() if comment.author else 'Anonyme'
    } for comment in comments]
    
    return JsonResponse({
        'comments': comments_data,
        'participant_name': f"{participant.user.first_name} {participant.user.last_name}"
    })
@login_required
def mark_notification_read(request, notification_id):
    """Marquer une notification comme lue."""
    try:
        # Récupérer la notification
        notification = Notification.objects.get(
            id=notification_id,
            user=request.user
        )
        
        # Marquer comme lue
        notification.is_read = True
        notification.save()
        
        # Retourner au bon endroit selon le type de notification
        if notification.related_object_type == 'Session':
            return redirect('core:session_detail', session_id=notification.related_object_id)
        
        # Par défaut, rediriger vers la liste des notifications
        return redirect('core:notifications_list')
        
    except Notification.DoesNotExist:
        messages.error(request, 'Notification introuvable.')
        return redirect('core:notifications_list')